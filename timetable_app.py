"""
================================================================================
中学校時間割自動生成アプリ  timetable_app.py
================================================================================

【実行方法】
1. ライブラリをインストール:
   pip install streamlit openpyxl ortools pandas

2. アプリを起動:
   streamlit run timetable_app.py

3. ブラウザで http://localhost:8501 を開く

【動作環境】
  Python 3.10 以上 / streamlit >= 1.28 / openpyxl >= 3.1
  ortools >= 9.0 / pandas >= 1.5

================================================================================
"""

# ============================================================
# インポート
# ============================================================
import streamlit as st
import openpyxl
from openpyxl.styles import PatternFill, Font, Border, Side, Alignment
from openpyxl.utils import get_column_letter
import io
import pandas as pd
from datetime import datetime
import time
import copy

# ============================================================
# ① 固定定数（変更禁止）
# ============================================================

SUBJECTS = [
    "国語", "数学", "社会", "理科", "英語",
    "保体", "技術", "家庭", "美術", "音楽",
    "学活", "総合", "道徳", "その他１", "その他２"
]

DAYS   = ["月", "火", "水", "木", "金"]
GRADES = [1, 2, 3]

COMMON_SUBJECTS = ["技術", "家庭", "保体", "美術", "音楽"]

COLOR_HEADER = "D9D9D9"
COLOR_MANUAL = "CCE5FF"
COLOR_AUTO   = "CCFFCC"
COLOR_SG     = "FFE5CC"
COLOR_ABSENT = "C0C0C0"

STEP_LABELS = {
    1: "STEP 1: クラス数設定",
    2: "STEP 2: 教員登録",
    3: "STEP 3: コマ数設定",
    4: "STEP 4: 週コマ数設定",
    5: "STEP 5: 担当教員割り当て",
    6: "STEP 6: 不在コマ設定",
    7: "STEP 7: 特別教室設定",
    8: "STEP 8: 手動時間割入力",
    9: "STEP 9: 少人数学級設定",
    10: "STEP 10: クラス間教科同期",
    0: "🎲 時間割生成・出力",
}

# ============================================================
# ② session_state 全キー初期化
# ============================================================

def init_session_state():
    defaults = {
        "grade_classes":          {g: 1 for g in GRADES},
        "teachers":               [],
        "periods_per_day":        {d: 6 for d in DAYS},
        "weekly_periods":         {},
        "assignments":            {},
        "unavailable":            {},
        "special_rooms":          [],
        "manual_timetable":       {},
        "manual_timetable_sg":    {},
        "small_group_classes":    {},
        "class_subject_sync":     [],
        "soft_grade_grouping":    True,
        "soft_priority_subjects": [],
        "soft_first_subjects":    [],
        "generated_timetable":    {},
        "generated_timetable_sg": {},
        "timetable_history":      [],   # 過去の生成結果を保持するリスト
        "swap_undo_stack":        [],   # 手動編集の undo スタック
        "swap2_slot_a":           None, # 2コマ入れ替え: 選択スロットA
        "swap2_slot_b":           None, # 2コマ入れ替え: 選択スロットB
        "swap2_result":           None, # 2コマ入れ替え: 確認結果
        "swap2_last_teacher":     None, # 2コマ入れ替え: 直前の選択教員
        "edit_cls_slot_a":        None, # 手動編集/クラスから選ぶ: 選択スロットA(day,period,subj)
        "edit_cls_slot_b":        None, # 手動編集/クラスから選ぶ: 選択スロットB(day,period,subj)
        "edit_cls_last":          None, # 手動編集/クラスから選ぶ: 直前の選択クラス
        "edit_cls_cands":         None, # 手動編集/クラスから選ぶ: 候補リスト(キャッシュ)
        "edit2_slot_a":           None, # 手動編集/教員から選ぶ: 2コマ入れ替えスロットA
        "edit2_slot_b":           None, # 手動編集/教員から選ぶ: 2コマ入れ替えスロットB
        "edit2_result":           None, # 手動編集/教員から選ぶ: 2コマ入れ替え確認結果
        "edit2_last_teacher":     None, # 手動編集/教員から選ぶ: 直前の選択教員
        "edit2_cands":            None, # 手動編集/教員から選ぶ: スロットAの候補キャッシュ
        "current_step":           1,
        "timetable_pattern_no":   0,
    }
    for key, val in defaults.items():
        if key not in st.session_state:
            st.session_state[key] = val

# ============================================================
# ③ ヘルパー関数
# ============================================================

def get_all_classes() -> list[str]:
    result = []
    for g in GRADES:
        count = st.session_state["grade_classes"].get(g, 0)
        for c in range(1, count + 1):
            result.append(f"{g}年{c}組")
    return result

def get_grade(class_name: str) -> int:
    return int(class_name[0])

def get_class_index(class_name: str) -> int:
    classes = get_all_classes()
    return classes.index(class_name) if class_name in classes else 999

def get_total_periods_per_week() -> int:
    return sum(st.session_state["periods_per_day"].values())

def get_max_periods() -> int:
    return max(st.session_state["periods_per_day"].values())

def get_teachers_for_slot(class_name: str, subject: str) -> list[str]:
    raw = st.session_state["assignments"].get(class_name, {}).get(subject, [])
    # 万一 ["山八,金場"] のようにカンマ区切りが1要素リストに入っていた場合も正しく展開する
    result = []
    for item in raw:
        for t in str(item).split(","):
            t = t.strip()
            if t:
                result.append(t)
    return result

def is_special_room_subject(subject: str) -> tuple[bool, int]:
    for room in st.session_state["special_rooms"]:
        if room["subject"] == subject:
            return True, room["capacity"]
    return False, 999

def ensure_class_keys():
    cls_list = get_all_classes()
    wp   = st.session_state["weekly_periods"]
    asgn = st.session_state["assignments"]
    for cls in cls_list:
        wp.setdefault(cls, {s: 0 for s in SUBJECTS})
        asgn.setdefault(cls, {s: [] for s in SUBJECTS})
    for cls in list(wp.keys()):
        if cls not in cls_list:
            del wp[cls]
    for cls in list(asgn.keys()):
        if cls not in cls_list:
            del asgn[cls]

# ============================================================
# ④ Excel 共通書式ヘルパー
# ============================================================

def _xl_fill(color: str) -> PatternFill:
    return PatternFill("solid", fgColor=color)

def _xl_border() -> Border:
    s = Side(style="thin")
    return Border(left=s, right=s, top=s, bottom=s)

def _xl_align(wrap: bool = True) -> Alignment:
    return Alignment(horizontal="center", vertical="center", wrap_text=wrap)

def _xl_write(ws, row: int, col: int, value,
              color: str | None = None, bold: bool = False,
              width: int | None = None):
    cell = ws.cell(row=row, column=col, value=value)
    cell.border    = _xl_border()
    cell.alignment = _xl_align()
    if color:
        cell.fill = _xl_fill(color)
    if bold:
        cell.font = Font(bold=True)
    if width:
        ws.column_dimensions[get_column_letter(col)].width = width
    return cell

# ============================================================
# Part 2: Excel I/O 補助スタイル関数
# ============================================================

def _apply_header_style(ws, row_num: int = 1):
    """指定行にヘッダースタイル（太字・薄グレー・罫線）を適用する。"""
    fill   = PatternFill("solid", fgColor=COLOR_HEADER)
    font   = Font(bold=True)
    thin   = Side(style="thin")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    for cell in ws[row_num]:
        cell.fill      = fill
        cell.font      = font
        cell.border    = border
        cell.alignment = Alignment(horizontal="center", vertical="center")


def _set_col_widths(ws, width: int = 14):
    """全列の幅を設定する。"""
    for col in ws.columns:
        ws.column_dimensions[col[0].column_letter].width = width


def _apply_cell_border(ws):
    """全データセルに罫線を適用する。"""
    thin   = Side(style="thin")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    for row in ws.iter_rows():
        for cell in row:
            cell.border = border


# ============================================================
# Part 2: Excel関数① テンプレート生成
# ============================================================

@st.cache_data
def generate_template_excel() -> bytes:
    """
    全10シートのテンプレートExcelをBytesIOで返す。
    サイドバーのダウンロードボタンに渡す。
    内容は固定なので st.cache_data でキャッシュする。
    """
    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    # ── シート1: クラス設定 ──────────────────────────────────
    ws = wb.create_sheet("クラス設定")
    ws.append(["学年", "クラス数"])
    for g in GRADES:
        ws.append([f"{g}年", 1])
    _apply_header_style(ws)
    _set_col_widths(ws, 12)

    # ── シート2: 先生リスト (ヘッダーなし: A=先生名) ─────────────
    ws = wb.create_sheet("先生リスト")
    ws.append(["（例）古澤"])
    _set_col_widths(ws, 16)

    # ── シート2b: 教科リスト (ヘッダーなし: A=教科名, B=週コマ数, C=必修フラグ) ──
    ws = wb.create_sheet("教科リスト")
    for subj in SUBJECTS:
        ws.append([subj, 0, 1])
    _set_col_widths(ws, 14)

    # ── シート3: コマ数設定 ──────────────────────────────────
    ws = wb.create_sheet("コマ数設定")
    ws.append(["曜日", "1日のコマ数"])
    for d in DAYS:
        ws.append([d, 6])
    _apply_header_style(ws)
    _set_col_widths(ws, 14)

    # ── シート4: 週コマ数 ────────────────────────────────────
    ws = wb.create_sheet("週コマ数")
    ws.append(["学年クラス"] + SUBJECTS)
    ws.append(["（例）1年1組"] + [0] * len(SUBJECTS))
    _apply_header_style(ws)
    _set_col_widths(ws, 10)
    ws.column_dimensions["A"].width = 14

    # ── シート5: 担当割り当て ────────────────────────────────
    ws = wb.create_sheet("担当割り当て")
    ws.append(["学年クラス"] + SUBJECTS)
    ws.append(["（例）1年1組"] + [""] * len(SUBJECTS))
    ws.cell(row=3, column=1).value = "※複数教員はカンマ区切り（例: 古澤,田中）"
    _apply_header_style(ws)
    _set_col_widths(ws, 10)
    ws.column_dimensions["A"].width = 14

    # ── シート6: 不在コマ ────────────────────────────────────
    ws = wb.create_sheet("不在コマ")
    ws.append(["教員名", "曜日", "時限"])
    ws.append(["（例）古澤", "月", 4])
    ws.append(["（例）古澤", "水", 5])
    _apply_header_style(ws)
    _set_col_widths(ws, 14)

    # ── シート7: 特別教室 ────────────────────────────────────
    ws = wb.create_sheet("特別教室")
    ws.append(["教室名", "対応教科", "同時使用可能クラス数"])
    ws.append(["理科室", "理科", 1])
    ws.append(["体育館", "保体", 1])
    ws.append(["音楽室", "音楽", 1])
    _apply_header_style(ws)
    _set_col_widths(ws, 18)

    # ── シート8: 手動時間割 ──────────────────────────────────
    ws = wb.create_sheet("手動時間割")
    ws.append(["学年クラス", "曜日", "時限", "教科"])
    ws.append(["（例）1年1組", "月", 1, "国語"])
    _apply_header_style(ws)
    _set_col_widths(ws, 14)

    # ── シート9: 少人数学級設定 ──────────────────────────────
    ws = wb.create_sheet("少人数学級設定")
    ws.append([
        "少人数学級名",
        "所属クラス（カンマ区切り）",
        "担当教員（カンマ区切り）",
        "同期グループ番号",
        "グループ内クラス（カンマ区切り）"
    ])
    ws.append(["少人数A", "1年1組,2年3組,3年1組", "田中", 1, "1年1組,3年1組"])
    ws.append(["少人数A", "", "", 2, "2年3組"])
    ws.cell(row=4, column=1).value = "※同期グループ未設定なら列D・Eは空欄"
    _apply_header_style(ws)
    _set_col_widths(ws, 22)

    # ── シート10: 少人数手動時間割 ───────────────────────────
    ws = wb.create_sheet("少人数手動時間割")
    ws.append(["少人数学級名", "曜日", "時限", "教科"])
    ws.append(["（例）少人数A", "月", 2, "音楽"])
    _apply_header_style(ws)
    _set_col_widths(ws, 16)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ============================================================
# Part 1b: 入力正規化ヘルパー
# ============================================================
import unicodedata as _ud

def normalize_value(val) -> str:
    """全角→半角 + strip"""
    if val is None:
        return ""
    return _ud.normalize("NFKC", str(val)).strip()

def normalize_str(val) -> str:
    """normalize_value の別名（互換性）"""
    return normalize_value(val)

def safe_int(val, default: int = 0) -> int:
    """正規化して整数に変換。失敗時は default を返す"""
    try:
        return int(normalize_value(val))
    except (ValueError, TypeError):
        return default

# ============================================================
# Part 2: Excel関数② 設定ファイル読み込み
# ============================================================

def load_settings_from_excel(uploaded_file):
    """
    Excelファイルを読み込み session_state を全て上書きする。
    シートが存在しない場合はそのSTEPをスキップし警告を表示する。
    """
    try:
        wb = openpyxl.load_workbook(uploaded_file, data_only=True)

        # ── シート1: クラス設定 ──────────────────────────────
        if "クラス設定" in wb.sheetnames:
            ws = wb["クラス設定"]
            for row in ws.iter_rows(min_row=2, values_only=True):
                if row[0] and row[1]:
                    try:
                        g = int(str(row[0]).replace("年", ""))
                        st.session_state["grade_classes"][g] = int(row[1])
                    except (ValueError, TypeError):
                        pass

        # ── シート2: 先生リスト (ヘッダーなし, row1から) / 旧: 教員リスト 両対応 ──
        _teacher_sheet = None
        if "先生リスト" in wb.sheetnames:
            _teacher_sheet = wb["先生リスト"]
        elif "教員リスト" in wb.sheetnames:
            _teacher_sheet = wb["教員リスト"]
        if _teacher_sheet is not None:
            teachers = []
            # 先生リストはヘッダーなし (row1 = データ先頭)
            # 教員リストはヘッダーあり (row1 = ヘッダー) → どちらも全行チェック
            for row in _teacher_sheet.iter_rows(min_row=1, values_only=True):
                val = row[0] if row else None
                if (val and str(val).strip()
                        and str(val).strip() not in ("教員名", "先生名")
                        and not str(val).startswith("（例）")):
                    teachers.append(normalize_str(str(val)))
            st.session_state["teachers"] = teachers

        # ── シート2b: 教科リスト (ヘッダーなし: A=教科名, B=週コマ数, C=必修フラグ) ──
        if "教科リスト" in wb.sheetnames:
            ws = wb["教科リスト"]
            subject_defaults = {}
            for row in ws.iter_rows(min_row=1, values_only=True):
                if not row or not row[0]:
                    continue
                subj = normalize_str(str(row[0]))
                if subj not in SUBJECTS:
                    continue
                weekly_cnt = safe_int(row[1]) if len(row) > 1 else 0
                # required_flag = row[2] if len(row) > 2 else 1  # 将来利用可
                subject_defaults[subj] = weekly_cnt
            # 教科リストの週コマ数を全クラスのデフォルトとして適用
            if subject_defaults:
                classes = [f"{g}年{c}組"
                           for g in GRADES
                           for c in range(1, st.session_state["grade_classes"].get(g, 0) + 1)]
                if classes:
                    wp = st.session_state.get("weekly_periods", {})
                    for cls in classes:
                        if cls not in wp:
                            wp[cls] = {}
                        for subj, cnt in subject_defaults.items():
                            if wp[cls].get(subj, 0) == 0:
                                wp[cls][subj] = cnt
                    st.session_state["weekly_periods"] = wp

        # ── シート3: コマ数設定 ──────────────────────────────
        if "コマ数設定" in wb.sheetnames:
            ws = wb["コマ数設定"]
            for row in ws.iter_rows(min_row=2, values_only=True):
                if row[0] and row[1] and row[0] in DAYS:
                    st.session_state["periods_per_day"][row[0]] = int(row[1])

        # ── シート4: 週コマ数 ────────────────────────────────
        if "週コマ数" in wb.sheetnames:
            ws = wb["週コマ数"]
            headers = [c.value for c in ws[1]][1:]
            weekly = {}
            for row in ws.iter_rows(min_row=2, values_only=True):
                cls = row[0]
                if not cls or str(cls).startswith("（例）"):
                    continue
                cls = str(cls).strip()
                weekly[cls] = {}
                for i, subj in enumerate(headers):
                    if subj and i + 1 < len(row):
                        weekly[cls][subj] = int(row[i + 1] or 0)
            st.session_state["weekly_periods"] = weekly

        # ── シート5: 担当割り当て ────────────────────────────
        if "担当割り当て" in wb.sheetnames:
            ws = wb["担当割り当て"]
            headers = [c.value for c in ws[1]][1:]
            assignments = {}
            for row in ws.iter_rows(min_row=2, values_only=True):
                cls = row[0]
                if not cls or str(cls).startswith("（例）") \
                        or str(cls).startswith("※"):
                    continue
                cls = str(cls).strip()
                assignments[cls] = {}
                for i, subj in enumerate(headers):
                    if subj and i + 1 < len(row) and row[i + 1]:
                        teachers = [t.strip()
                                    for t in str(row[i + 1]).split(",")
                                    if t.strip()]
                        assignments[cls][subj] = teachers
                    else:
                        assignments[cls][subj] = []
            st.session_state["assignments"] = assignments

        # ── シート6: 不在コマ ────────────────────────────────
        if "不在コマ" in wb.sheetnames:
            ws = wb["不在コマ"]
            unavail = {}
            for row in ws.iter_rows(min_row=2, values_only=True):
                teacher, day, period = row[0], row[1], row[2]
                if not teacher or str(teacher).startswith("（例）"):
                    continue
                teacher = str(teacher).strip()
                if day in DAYS and period:
                    unavail.setdefault(teacher, {}).setdefault(
                        str(day), []).append(int(period))
            st.session_state["unavailable"] = unavail

        # ── シート7: 特別教室 ────────────────────────────────
        if "特別教室" in wb.sheetnames:
            ws = wb["特別教室"]
            rooms = []
            for row in ws.iter_rows(min_row=2, values_only=True):
                name, subj, cap = row[0], row[1], row[2]
                if not name or str(name).startswith("（例）"):
                    continue
                if name and subj and cap:
                    rooms.append({
                        "name":     str(name).strip(),
                        "subject":  str(subj).strip(),
                        "capacity": int(cap)
                    })
            st.session_state["special_rooms"] = rooms

        # ── シート8: 手動時間割 ──────────────────────────────
        if "手動時間割" in wb.sheetnames:
            ws = wb["手動時間割"]
            manual = {}
            for row in ws.iter_rows(min_row=2, values_only=True):
                cls, day, period, subj = row[0], row[1], row[2], row[3]
                if not cls or str(cls).startswith("（例）"):
                    continue
                if cls and day in DAYS and period and subj:
                    cls = str(cls).strip()
                    manual.setdefault(cls, {}).setdefault(
                        str(day), {})[int(period)] = str(subj).strip()
            st.session_state["manual_timetable"] = manual

        # ── シート9: 少人数学級設定 ──────────────────────────
        if "少人数学級設定" in wb.sheetnames:
            ws = wb["少人数学級設定"]
            sg = {}
            for row in ws.iter_rows(min_row=2, values_only=True):
                sg_name = row[0]
                if sg_name and not str(sg_name).startswith("（例）") \
                        and not str(sg_name).startswith("※"):
                    sg_name = str(sg_name).strip()
                    if sg_name not in sg:
                        sg[sg_name] = {
                            "classes":        [normalize_str(c)
                                               for c in str(row[1]).split(",")
                                               if c.strip()] if row[1] else [],
                            "teachers":       [t.strip()
                                               for t in str(row[2]).split(",")
                                               if t.strip()] if row[2] else [],
                            "weekly_periods": {},
                            "sync_groups":    []
                        }
                    if row[3] and row[4]:
                        group_classes = [c.strip()
                                         for c in str(row[4]).split(",")
                                         if c.strip()]
                        group_no = int(row[3])
                        while len(sg[sg_name]["sync_groups"]) < group_no:
                            sg[sg_name]["sync_groups"].append([])
                        sg[sg_name]["sync_groups"][group_no - 1] = group_classes
                elif not sg_name and sg:
                    last_sg = list(sg.keys())[-1]
                    if row[3] and row[4]:
                        group_classes = [c.strip()
                                         for c in str(row[4]).split(",")
                                         if c.strip()]
                        group_no = int(row[3])
                        while len(sg[last_sg]["sync_groups"]) < group_no:
                            sg[last_sg]["sync_groups"].append([])
                        sg[last_sg]["sync_groups"][group_no - 1] = group_classes
            st.session_state["small_group_classes"] = sg
            # ウィジェットキーを強制上書き（Excelロード後の古いキャッシュを排除）
            all_valid_classes = get_all_classes()
            for sg_name_, sg_data_ in sg.items():
                st.session_state[f"sg_cls_{sg_name_}"] = [
                    c for c in sg_data_.get("classes", [])
                    if c in all_valid_classes
                ]
                st.session_state[f"sg_teacher_{sg_name_}"] = sg_data_.get("teachers", [])
                for subj_ in COMMON_SUBJECTS:
                    st.session_state[f"sg_wp_{sg_name_}_{subj_}"] = \
                        sg_data_.get("weekly_periods", {}).get(subj_, 0)
                for gi_, grp_ in enumerate(sg_data_.get("sync_groups", [])):
                    st.session_state[f"sg_group_{sg_name_}_{gi_}"] = grp_

        # ── シート10: 少人数手動時間割 ───────────────────────
        if "少人数手動時間割" in wb.sheetnames:
            ws = wb["少人数手動時間割"]
            manual_sg = {}
            for row in ws.iter_rows(min_row=2, values_only=True):
                sg_name, day, period, subj = row[0], row[1], row[2], row[3]
                if not sg_name or str(sg_name).startswith("（例）"):
                    continue
                if sg_name and day in DAYS and period and subj:
                    sg_name = str(sg_name).strip()
                    manual_sg.setdefault(sg_name, {}).setdefault(
                        str(day), {})[int(period)] = str(subj).strip()
            st.session_state["manual_timetable_sg"] = manual_sg

        # ── ウィジェットキーをクリアして各STEPの入力欄を再初期化 ──
        # Streamlit はウィジェットキーが session_state に残っていると
        # value= パラメータより優先するため、Excel ロード後に古いキーを
        # 削除しておかないと読み込んだ値が画面に反映されない。
        # ※ "grade_" は "grade_classes" を誤って削除しないよう、
        #   数字サフィックスのもの（grade_1, grade_2, grade_3）のみ対象にする。
        for _k in list(st.session_state.keys()):
            if _k.startswith("grade_") and _k[6:].isdigit():
                del st.session_state[_k]
            elif _k.startswith(("ppd_", "wp_", "assign_", "unavail_")):
                del st.session_state[_k]

        st.success("✅ Excelファイルから設定を読み込みました（全STEP上書き）")

    except Exception as e:
        st.error(f"❌ 読み込みエラー: {e}")
        st.info("💡 テンプレートExcelを使用して入力してください")


# ============================================================
# Part 2: Excel関数③ 設定ファイル書き出し
# ============================================================

def save_settings_to_excel() -> bytes:
    """
    現在のsession_stateをExcelに書き出してBytesIOで返す。
    generate_template_excel()と同じシート構成・書式で出力する。
    """
    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    # シート1: クラス設定
    ws = wb.create_sheet("クラス設定")
    ws.append(["学年", "クラス数"])
    for g in GRADES:
        ws.append([f"{g}年", st.session_state["grade_classes"].get(g, 0)])
    _apply_header_style(ws); _set_col_widths(ws, 12)

    # シート2: 先生リスト (ヘッダーなし)
    ws = wb.create_sheet("先生リスト")
    for t in st.session_state["teachers"]:
        ws.append([t])
    _set_col_widths(ws, 16)

    # シート2b: 教科リスト (ヘッダーなし)
    ws = wb.create_sheet("教科リスト")
    weekly_p = st.session_state.get("weekly_periods", {})
    classes = [f"{g}年{c}組"
               for g in GRADES
               for c in range(1, st.session_state["grade_classes"].get(g, 0) + 1)]
    first_cls = classes[0] if classes else None
    for subj in SUBJECTS:
        cnt = weekly_p.get(first_cls, {}).get(subj, 0) if first_cls else 0
        ws.append([subj, cnt, 1])
    _set_col_widths(ws, 14)

    # シート3: コマ数設定
    ws = wb.create_sheet("コマ数設定")
    ws.append(["曜日", "1日のコマ数"])
    for d in DAYS:
        ws.append([d, st.session_state["periods_per_day"].get(d, 6)])
    _apply_header_style(ws); _set_col_widths(ws, 14)

    # シート4: 週コマ数
    ws = wb.create_sheet("週コマ数")
    ws.append(["学年クラス"] + SUBJECTS)
    for cls in get_all_classes():
        row = [cls]
        for subj in SUBJECTS:
            row.append(
                st.session_state["weekly_periods"].get(cls, {}).get(subj, 0))
        ws.append(row)
    _apply_header_style(ws); _set_col_widths(ws, 10)
    ws.column_dimensions["A"].width = 14

    # シート5: 担当割り当て
    ws = wb.create_sheet("担当割り当て")
    ws.append(["学年クラス"] + SUBJECTS)
    for cls in get_all_classes():
        row = [cls]
        for subj in SUBJECTS:
            teachers = st.session_state["assignments"].get(
                cls, {}).get(subj, [])
            row.append(",".join(teachers))
        ws.append(row)
    _apply_header_style(ws); _set_col_widths(ws, 10)
    ws.column_dimensions["A"].width = 14

    # シート6: 不在コマ
    ws = wb.create_sheet("不在コマ")
    ws.append(["教員名", "曜日", "時限"])
    for teacher, day_map in st.session_state["unavailable"].items():
        for day, periods in day_map.items():
            for p in periods:
                ws.append([teacher, day, p])
    _apply_header_style(ws); _set_col_widths(ws, 14)

    # シート7: 特別教室
    ws = wb.create_sheet("特別教室")
    ws.append(["教室名", "対応教科", "同時使用可能クラス数"])
    for room in st.session_state["special_rooms"]:
        ws.append([room["name"], room["subject"], room["capacity"]])
    _apply_header_style(ws); _set_col_widths(ws, 18)

    # シート8: 手動時間割
    ws = wb.create_sheet("手動時間割")
    ws.append(["学年クラス", "曜日", "時限", "教科"])
    for cls, day_map in st.session_state["manual_timetable"].items():
        for day, period_map in day_map.items():
            for period, subj in period_map.items():
                ws.append([cls, day, period, subj])
    _apply_header_style(ws); _set_col_widths(ws, 14)

    # シート9: 少人数学級設定
    ws = wb.create_sheet("少人数学級設定")
    ws.append(["少人数学級名", "所属クラス（カンマ区切り）",
               "担当教員（カンマ区切り）", "同期グループ番号",
               "グループ内クラス（カンマ区切り）"])
    for sg_name, sg_data in st.session_state["small_group_classes"].items():
        sync_groups = sg_data.get("sync_groups", [])
        if sync_groups:
            for gi, group in enumerate(sync_groups):
                if gi == 0:
                    ws.append([
                        sg_name,
                        ",".join(sg_data.get("classes", [])),
                        ",".join(sg_data.get("teachers", [])),
                        gi + 1,
                        ",".join(group)
                    ])
                else:
                    ws.append(["", "", "", gi + 1, ",".join(group)])
        else:
            ws.append([
                sg_name,
                ",".join(sg_data.get("classes", [])),
                ",".join(sg_data.get("teachers", [])),
                "", ""
            ])
    _apply_header_style(ws); _set_col_widths(ws, 22)

    # シート10: 少人数手動時間割
    ws = wb.create_sheet("少人数手動時間割")
    ws.append(["少人数学級名", "曜日", "時限", "教科"])
    for sg_name, day_map in st.session_state["manual_timetable_sg"].items():
        for day, period_map in day_map.items():
            for period, subj in period_map.items():
                ws.append([sg_name, day, period, subj])
    _apply_header_style(ws); _set_col_widths(ws, 16)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ============================================================
# Part 2: サイドバー UI
# ============================================================

def render_sidebar():
    """
    サイドバーにExcelボタン群・STEPナビゲーションを表示する。
    """
    st.sidebar.title("📅 中学校時間割生成")
    st.sidebar.markdown("---")

    # Excelテンプレートダウンロード
    st.sidebar.download_button(
        label="📥 テンプレートDL",
        data=generate_template_excel(),
        file_name="時間割テンプレート.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        help="入力用テンプレートをダウンロードします"
    )

    # 設定ファイルアップロード
    uploaded = st.sidebar.file_uploader(
        "📤 設定ファイルをアップロード",
        type=["xlsx"],
        help="保存済みの設定ファイルをアップロードします（全STEP上書き）"
    )
    if uploaded:
        # ファイルの同一性を名前+サイズで判定し、新規アップロード時のみ読み込む
        # （画面遷移のたびに再読み込みされて手動入力が消えるのを防ぐ）
        file_key = f"{uploaded.name}_{uploaded.size}"
        if st.session_state.get("_last_uploaded_excel_key") != file_key:
            load_settings_from_excel(uploaded)
            st.session_state["_last_uploaded_excel_key"] = file_key

    # 現在設定を保存
    st.sidebar.download_button(
        label="💾 現在の設定を保存",
        data=save_settings_to_excel(),
        file_name=f"時間割設定_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        help="現在の設定をExcelに保存します"
    )

    st.sidebar.markdown("---")
    st.sidebar.markdown("### 📋 設定STEP")

    # STEPナビゲーションボタン
    step_labels = {
        1: "① クラス数設定",
        2: "② 教員登録",
        3: "③ コマ数設定",
        4: "④ 週コマ数",
        5: "⑤ 担当教員割り当て",
        6: "⑥ 不在コマ設定",
        7: "⑦ 特別教室設定",
        8: "⑧ 手動時間割入力",
        9: "⑨ 少人数学級設定",
        10: "⑩ クラス間教科同期",
    }
    for step_num, label in step_labels.items():
        if st.sidebar.button(label, key=f"nav_{step_num}",
                             use_container_width=True):
            st.session_state["current_step"] = step_num
            st.rerun()

    st.sidebar.markdown("---")
    if st.sidebar.button("🚀 時間割を生成・確認",
                         type="primary", use_container_width=True):
        st.session_state["current_step"] = 0
        st.rerun()


# ============================================================
# Part 2: STEP 1〜5 UI
# ============================================================

def render_step1():
    st.header("STEP 1｜学年・クラス数設定")
    st.caption("各学年のクラス数を設定してください（1〜6クラス）")

    cols = st.columns(3)
    for i, g in enumerate(GRADES):
        with cols[i]:
            st.session_state["grade_classes"][g] = st.number_input(
                f"{g}年生のクラス数",
                min_value=1, max_value=6,
                value=st.session_state["grade_classes"].get(g, 1),
                step=1,
                key=f"grade_{g}"
            )

    classes = get_all_classes()
    st.markdown("---")
    st.info(f"📋 設定クラス一覧（全 **{len(classes)}** クラス）\n\n"
            + "　".join(classes))


def render_step2():
    st.header("STEP 2｜教員登録")
    st.caption("時間割に関わる教員を全員登録してください")

    col1, col2 = st.columns([3, 1])
    with col1:
        new_name = st.text_input(
            "教員名を入力",
            placeholder="例: 古澤",
            key="new_teacher_input",
            label_visibility="collapsed"
        )
    with col2:
        if st.button("➕ 追加", use_container_width=True):
            name = new_name.strip()
            if name and name not in st.session_state["teachers"]:
                st.session_state["teachers"].append(name)
                st.rerun()
            elif name in st.session_state["teachers"]:
                st.warning("同じ名前の教員がすでに登録されています")

    st.markdown("---")
    st.markdown(f"#### 登録済み教員（{len(st.session_state['teachers'])}名）")

    if not st.session_state["teachers"]:
        st.info("教員が登録されていません")
    else:
        for i, teacher in enumerate(st.session_state["teachers"]):
            col1, col2 = st.columns([5, 1])
            col1.write(f"👤 {teacher}")
            if col2.button("🗑️ 削除", key=f"del_teacher_{i}",
                           use_container_width=True):
                st.session_state["teachers"].pop(i)
                st.rerun()


def render_step3():
    st.header("STEP 3｜曜日別1日コマ数設定")
    st.caption("各曜日の1日のコマ数を設定してください（1〜8コマ）")

    cols = st.columns(5)
    for i, d in enumerate(DAYS):
        with cols[i]:
            st.session_state["periods_per_day"][d] = st.number_input(
                f"{d}曜日",
                min_value=1, max_value=8,
                value=st.session_state["periods_per_day"].get(d, 6),
                step=1,
                key=f"ppd_{d}"
            )

    total = get_total_periods_per_week()
    st.markdown("---")
    st.info(f"📊 週合計コマ数: **{total}** コマ")


def render_step4():
    st.header("STEP 4｜週コマ数設定")
    st.caption("各クラス・各教科の週あたりのコマ数を入力してください")

    classes = get_all_classes()
    total_per_week = get_total_periods_per_week()

    if not classes:
        st.warning("先にSTEP 1でクラス数を設定してください")
        return

    for cls in classes:
        with st.expander(f"📘 {cls}", expanded=False):
            if cls not in st.session_state["weekly_periods"]:
                st.session_state["weekly_periods"][cls] = {s: 0 for s in SUBJECTS}

            total = 0
            cols_per_row = 5
            for row_start in range(0, len(SUBJECTS), cols_per_row):
                row_subjects = SUBJECTS[row_start:row_start + cols_per_row]
                cols = st.columns(cols_per_row)
                for j, subj in enumerate(row_subjects):
                    with cols[j]:
                        val = st.number_input(
                            subj,
                            min_value=0,
                            max_value=total_per_week,
                            value=st.session_state["weekly_periods"][cls].get(subj, 0),
                            step=1,
                            key=f"wp_{cls}_{subj}"
                        )
                        st.session_state["weekly_periods"][cls][subj] = val
                        total += val

            if total > total_per_week:
                st.error(
                    f"⚠️ 合計 {total} コマ → 週上限 {total_per_week} コマを超過！")
            elif total == total_per_week:
                st.success(f"✅ 合計 {total} コマ（週上限ちょうど）")
            else:
                st.warning(
                    f"📝 合計 {total} コマ（週上限まで残り {total_per_week - total} コマ）")


def render_step5():
    st.header("STEP 5｜教科担当教員割り当て")
    st.caption("各クラス・各教科の担当教員を選択してください（複数選択可）")

    classes  = get_all_classes()
    teachers = st.session_state["teachers"]

    if not teachers:
        st.warning("先にSTEP 2で教員を登録してください")
        return
    if not classes:
        st.warning("先にSTEP 1でクラス数を設定してください")
        return

    for cls in classes:
        with st.expander(f"📘 {cls}", expanded=False):
            if cls not in st.session_state["assignments"]:
                st.session_state["assignments"][cls] = {s: [] for s in SUBJECTS}

            cols_per_row = 5
            for row_start in range(0, len(SUBJECTS), cols_per_row):
                row_subjects = SUBJECTS[row_start:row_start + cols_per_row]
                cols = st.columns(cols_per_row)
                for j, subj in enumerate(row_subjects):
                    with cols[j]:
                        current = st.session_state["assignments"][cls].get(subj, [])
                        valid_current = [t for t in current if t in teachers]
                        selected = st.multiselect(
                            subj,
                            options=teachers,
                            default=valid_current,
                            key=f"assign_{cls}_{subj}"
                        )
                        st.session_state["assignments"][cls][subj] = selected


# ============================================================
# Part 3: STEP 6〜9 UI
# ============================================================

def render_step6():
    st.header("STEP 6｜教員不在コマ設定")
    st.caption("各教員が授業に入れない時限をチェックしてください")

    teachers = st.session_state["teachers"]
    if not teachers:
        st.warning("先にSTEP 2で教員を登録してください")
        return

    for teacher in teachers:
        with st.expander(f"👤 {teacher}", expanded=False):
            if teacher not in st.session_state["unavailable"]:
                st.session_state["unavailable"][teacher] = {}

            for day in DAYS:
                periods = st.session_state["periods_per_day"].get(day, 6)
                current_unavail = st.session_state["unavailable"][teacher].get(day, [])

                st.markdown(f"**{day}曜日**")
                cols = st.columns(periods)
                selected = []
                for p in range(1, periods + 1):
                    with cols[p - 1]:
                        checked = st.checkbox(
                            f"{p}限",
                            value=(p in current_unavail),
                            key=f"unavail_{teacher}_{day}_{p}"
                        )
                        if checked:
                            selected.append(p)
                st.session_state["unavailable"][teacher][day] = selected

            total_unavail = sum(
                len(v) for v in st.session_state["unavailable"][teacher].values()
            )
            if total_unavail > 0:
                st.caption(f"合計不在コマ数: {total_unavail} コマ/週")


def render_step7():
    st.header("STEP 7｜特別教室設定")
    st.caption("特別教室を登録し、対応教科と同時使用可能クラス数を設定してください")

    if st.button("➕ 特別教室を追加", key="add_room"):
        st.session_state["special_rooms"].append(
            {"name": "", "subject": SUBJECTS[0], "capacity": 1}
        )
        st.rerun()

    if not st.session_state["special_rooms"]:
        st.info("特別教室が登録されていません。上のボタンで追加してください。")
        return

    for i, room in enumerate(st.session_state["special_rooms"]):
        col1, col2, col3, col4 = st.columns([3, 2, 2, 1])
        with col1:
            room["name"] = st.text_input(
                "教室名",
                value=room.get("name", ""),
                placeholder="例: 理科室",
                key=f"room_name_{i}"
            )
        with col2:
            subj_idx = SUBJECTS.index(room["subject"]) \
                if room.get("subject") in SUBJECTS else 0
            room["subject"] = st.selectbox(
                "対応教科",
                options=SUBJECTS,
                index=subj_idx,
                key=f"room_subj_{i}"
            )
        with col3:
            room["capacity"] = st.number_input(
                "同時使用可能クラス数",
                min_value=1, max_value=10,
                value=room.get("capacity", 1),
                step=1,
                key=f"room_cap_{i}"
            )
        with col4:
            st.markdown("　")
            if st.button("🗑️", key=f"del_room_{i}", use_container_width=True):
                st.session_state["special_rooms"].pop(i)
                st.rerun()

    st.markdown("---")
    st.markdown("#### 設定中の特別教室")
    for room in st.session_state["special_rooms"]:
        if room.get("name"):
            st.write(
                f"🏫 **{room['name']}** → {room['subject']} "
                f"（同時 {room['capacity']} クラスまで）"
            )


def render_step8():
    st.header("STEP 8｜手動時間割入力")
    st.caption(
        "自動生成前に手動で配置したいコマを入力してください。"
        "手動入力したコマは自動生成で上書きされません（水色で表示）。"
    )

    tab_normal, tab_sg = st.tabs(["📘 通常学級", "📗 少人数学級"])

    # ── 通常学級タブ ─────────────────────────────────────────
    with tab_normal:
        classes = get_all_classes()
        if not classes:
            st.warning("先にSTEP 1でクラス数を設定してください")
        else:
            cls = st.selectbox("クラスを選択", classes, key="manual_cls_select")

            if cls not in st.session_state["manual_timetable"]:
                st.session_state["manual_timetable"][cls] = {}

            options = ["（未入力）"] + SUBJECTS

            for day in DAYS:
                periods = st.session_state["periods_per_day"].get(day, 6)
                st.markdown(f"**{day}曜日**")
                cols = st.columns(periods)
                day_data = st.session_state["manual_timetable"][cls].setdefault(day, {})

                for p in range(1, periods + 1):
                    with cols[p - 1]:
                        current_val = day_data.get(p, "（未入力）")
                        idx = options.index(current_val) \
                            if current_val in options else 0
                        chosen = st.selectbox(
                            f"{p}限",
                            options=options,
                            index=idx,
                            key=f"manual_{cls}_{day}_{p}"
                        )
                        if chosen == "（未入力）":
                            day_data.pop(p, None)
                        else:
                            day_data[p] = chosen

            total_manual = sum(
                len(v) for v in
                st.session_state["manual_timetable"].get(cls, {}).values()
            )
            st.markdown("---")
            col1, col2 = st.columns([3, 1])
            col1.info(f"📌 {cls} の手動入力コマ数: {total_manual} コマ")
            if col2.button(f"🗑️ {cls} をすべてクリア",
                           key=f"clear_manual_{cls}"):
                st.session_state["manual_timetable"][cls] = {}
                st.rerun()

    # ── 少人数学級タブ ───────────────────────────────────────
    with tab_sg:
        sg_list = list(st.session_state["small_group_classes"].keys())
        if not sg_list:
            st.info("先にSTEP 9で少人数学級を登録してください")
        else:
            sg_name = st.selectbox(
                "少人数学級を選択", sg_list, key="manual_sg_select"
            )

            if sg_name not in st.session_state["manual_timetable_sg"]:
                st.session_state["manual_timetable_sg"][sg_name] = {}

            options = ["（未入力）"] + SUBJECTS

            for day in DAYS:
                periods = st.session_state["periods_per_day"].get(day, 6)
                st.markdown(f"**{day}曜日**")
                cols = st.columns(periods)
                day_data = st.session_state["manual_timetable_sg"][sg_name].setdefault(day, {})

                for p in range(1, periods + 1):
                    with cols[p - 1]:
                        current_val = day_data.get(p, "（未入力）")
                        idx = options.index(current_val) \
                            if current_val in options else 0
                        chosen = st.selectbox(
                            f"{p}限",
                            options=options,
                            index=idx,
                            key=f"manual_sg_{sg_name}_{day}_{p}"
                        )
                        if chosen == "（未入力）":
                            day_data.pop(p, None)
                        else:
                            day_data[p] = chosen

            total_manual_sg = sum(
                len(v) for v in
                st.session_state["manual_timetable_sg"].get(sg_name, {}).values()
            )
            st.markdown("---")
            col1, col2 = st.columns([3, 1])
            col1.info(f"📌 {sg_name} の手動入力コマ数: {total_manual_sg} コマ")
            if col2.button(f"🗑️ {sg_name} をすべてクリア",
                           key=f"clear_manual_sg_{sg_name}"):
                st.session_state["manual_timetable_sg"][sg_name] = {}
                st.rerun()


def render_step9():
    st.header("STEP 9｜少人数学級設定")
    st.caption(
        "少人数学級を登録し、所属クラス（原籍学級）・担当教員・"
        "同期グループ（R6制約）を設定してください"
    )

    classes  = get_all_classes()
    teachers = st.session_state["teachers"]

    col1, col2 = st.columns([3, 1])
    with col1:
        new_sg = st.text_input(
            "少人数学級名を入力",
            placeholder="例: 少人数A、少人数理科1",
            key="new_sg_name_input",
            label_visibility="collapsed"
        )
    with col2:
        if st.button("➕ 追加", key="add_sg", use_container_width=True):
            name = new_sg.strip()
            if name and name not in st.session_state["small_group_classes"]:
                st.session_state["small_group_classes"][name] = {
                    "classes":        [],
                    "teachers":       [],
                    "weekly_periods": {s: 0 for s in COMMON_SUBJECTS},
                    "sync_groups":    []
                }
                st.rerun()
            elif name in st.session_state["small_group_classes"]:
                st.warning("同じ名前の少人数学級がすでに存在します")

    if not st.session_state["small_group_classes"]:
        st.info("少人数学級が登録されていません")
        return

    for sg_name in list(st.session_state["small_group_classes"].keys()):
        sg_data = st.session_state["small_group_classes"][sg_name]
        with st.expander(f"📗 {sg_name}", expanded=True):

            # ── 所属クラス設定 ───────────────────────────────
            # セッションステートキーを初期化（初回のみ）
            key_cls = f"sg_cls_{sg_name}"
            if key_cls not in st.session_state:
                st.session_state[key_cls] = [
                    c for c in sg_data.get("classes", []) if c in classes
                ]
            st.multiselect(
                "所属クラス（原籍学級）",
                options=classes,
                key=key_cls,
                help="この少人数学級に生徒が所属する原籍学級を全て選択"
            )
            st.session_state["small_group_classes"][sg_name]["classes"] = \
                st.session_state[key_cls]

            # ── 担当教員設定 ─────────────────────────────────
            key_teacher = f"sg_teacher_{sg_name}"
            if key_teacher not in st.session_state:
                st.session_state[key_teacher] = [
                    t for t in sg_data.get("teachers", []) if t in teachers
                ]
            st.multiselect(
                "担当教員",
                options=teachers,
                key=key_teacher,
                help="R1（教員重複禁止）の対象となります"
            )
            st.session_state["small_group_classes"][sg_name]["teachers"] = \
                st.session_state[key_teacher]

            # ── 週コマ数（共通教科のみ） ─────────────────────
            st.markdown("**週コマ数（共通教科）**")
            cols = st.columns(len(COMMON_SUBJECTS))
            if "weekly_periods" not in sg_data:
                st.session_state["small_group_classes"][sg_name]["weekly_periods"] = {}
            for ci, subj in enumerate(COMMON_SUBJECTS):
                with cols[ci]:
                    key_wp = f"sg_wp_{sg_name}_{subj}"
                    if key_wp not in st.session_state:
                        st.session_state[key_wp] = \
                            sg_data.get("weekly_periods", {}).get(subj, 0)
                    st.number_input(
                        subj,
                        min_value=0, max_value=10,
                        step=1,
                        key=key_wp
                    )
                    st.session_state["small_group_classes"][sg_name][
                        "weekly_periods"][subj] = st.session_state[key_wp]

            # ── 同期グループ設定（R6） ────────────────────────
            st.markdown("---")
            st.markdown("**同期グループ設定（R6: 少人数学級同期制約）**")
            st.caption(
                "同じグループ内のクラスは、同一時限に"
                "「全員が共通教科」または「全員が非共通教科」になるよう制約されます。"
                "グループ未設定の場合は所属クラス全員に一括適用されます。"
            )

            current_classes = st.session_state["small_group_classes"][sg_name]["classes"]

            if st.button("➕ グループを追加", key=f"add_group_{sg_name}"):
                st.session_state["small_group_classes"][sg_name]["sync_groups"].append([])
                st.rerun()

            current_groups = st.session_state["small_group_classes"][sg_name]["sync_groups"]

            if not current_groups:
                st.info(
                    "💡 グループ未設定: 所属クラス全員（"
                    + ", ".join(current_classes) + "）を一括で同期します"
                )
            else:
                for gi in range(len(current_groups)):
                    col1, col2 = st.columns([5, 1])
                    with col1:
                        valid_group = [c for c in current_groups[gi]
                                       if c in current_classes]
                        key_grp = f"sg_group_{sg_name}_{gi}"
                        if key_grp not in st.session_state:
                            st.session_state[key_grp] = valid_group
                        st.multiselect(
                            f"グループ {gi + 1}（2クラス以上選択）",
                            options=current_classes,
                            key=key_grp,
                            help="このグループ内で共通教科の時間帯を揃えます"
                        )
                        st.session_state["small_group_classes"][sg_name][
                            "sync_groups"][gi] = st.session_state[key_grp]
                    with col2:
                        st.markdown("　")
                        if st.button("🗑️", key=f"del_group_{sg_name}_{gi}",
                                     use_container_width=True):
                            st.session_state["small_group_classes"][sg_name][
                                "sync_groups"].pop(gi)
                            # グループキーをクリアして再初期化させる
                            for k in list(st.session_state.keys()):
                                if k.startswith(f"sg_group_{sg_name}_"):
                                    del st.session_state[k]
                            st.rerun()

                for gi, group in enumerate(current_groups):
                    if len(group) < 2:
                        st.warning(
                            f"⚠️ グループ {gi + 1}: "
                            "同期グループは2クラス以上必要です"
                        )

            # ── 少人数学級の削除 ─────────────────────────────
            st.markdown("---")
            if st.button(
                f"🗑️ 「{sg_name}」を削除",
                key=f"del_sg_{sg_name}",
                type="secondary"
            ):
                del st.session_state["small_group_classes"][sg_name]
                st.session_state["manual_timetable_sg"].pop(sg_name, None)
                # 関連するウィジェットキーをクリア
                for k in list(st.session_state.keys()):
                    if k.startswith(f"sg_cls_{sg_name}") or \
                       k.startswith(f"sg_teacher_{sg_name}") or \
                       k.startswith(f"sg_wp_{sg_name}_") or \
                       k.startswith(f"sg_group_{sg_name}_"):
                        del st.session_state[k]
                st.rerun()


# ============================================================
# STEP 10: クラス間教科同期設定
# ============================================================

def render_step10():
    st.header("STEP 10｜クラス間教科同期設定")
    st.caption(
        "異なるクラスの教科を同じ曜日・同じ時限に揃えるハード制約を設定します。\n\n"
        "例：「1年1組 技術」と「1年2組 家庭」を同期 → 1年1組が技術を行う時限に、"
        "1年2組も必ず家庭を行うよう拘束します（双方向）。"
    )

    pairs = st.session_state["class_subject_sync"]
    all_classes = get_all_classes()

    if not all_classes:
        st.warning("先にSTEP 1でクラスを設定してください")
        return

    # ── 新規ペア追加 ───────────────────────────────────────
    st.markdown("#### 同期ペアを追加")
    col1, col2, col3, col4, col5 = st.columns([2, 2, 2, 2, 1])
    with col1:
        new_c1 = st.selectbox("クラス①", all_classes, key="css_c1")
    with col2:
        new_s1 = st.selectbox("教科①", SUBJECTS, key="css_s1")
    with col3:
        new_c2 = st.selectbox("クラス②", all_classes, key="css_c2")
    with col4:
        new_s2 = st.selectbox("教科②", SUBJECTS, key="css_s2")
    with col5:
        st.markdown("<br>", unsafe_allow_html=True)
        if st.button("➕ 追加", key="css_add", use_container_width=True):
            if new_c1 == new_c2 and new_s1 == new_s2:
                st.warning("同じクラス・同じ教科のペアは追加できません")
            else:
                new_pair = {"class1": new_c1, "subject1": new_s1,
                            "class2": new_c2, "subject2": new_s2}
                reverse   = {"class1": new_c2, "subject1": new_s2,
                             "class2": new_c1, "subject2": new_s1}
                if new_pair in pairs or reverse in pairs:
                    st.warning("同じ組み合わせがすでに登録されています")
                else:
                    pairs.append(new_pair)
                    st.rerun()

    st.markdown("---")

    # ── 登録済みペア一覧 ───────────────────────────────────
    if not pairs:
        st.info("同期ペアが登録されていません")
        return

    st.markdown("#### 登録済み同期ペア")
    assignments = st.session_state.get("assignments", {})
    for i, pair in enumerate(pairs):
        c1, s1 = pair["class1"], pair["subject1"]
        c2, s2 = pair["class2"], pair["subject2"]
        col_a, col_b = st.columns([6, 1])
        with col_a:
            st.markdown(
                f"**{i+1}.** {c1}「{s1}」　↔　{c2}「{s2}」"
            )
            # 同一教員が両側に登録されていると R1 と矛盾するため警告
            teachers1 = set(assignments.get(c1, {}).get(s1, []))
            teachers2 = set(assignments.get(c2, {}).get(s2, []))
            conflict_teachers = teachers1 & teachers2
            if conflict_teachers:
                st.error(
                    f"⚠️ **担当教員の矛盾**: {', '.join(conflict_teachers)} 先生が "
                    f"{c1}「{s1}」と {c2}「{s2}」の両方に担当登録されています。\n\n"
                    "同期制約（R8）では「同じ時間帯に配置」が必要ですが、"
                    "教員重複禁止（R1）では「同じ時間帯に2クラス不可」のため矛盾が生じ、"
                    "時間割が組めなくなります。\n\n"
                    "**STEP 5 で担当教員を修正してください：**\n"
                    f"- {c1}「{s1}」→ {c1} 側を担当する先生のみ\n"
                    f"- {c2}「{s2}」→ {c2} 側を担当する先生のみ"
                )
        with col_b:
            if st.button("🗑️", key=f"css_del_{i}", help="削除"):
                pairs.pop(i)
                st.rerun()


# ============================================================
# Part 3: ソフト制約設定UI
# ============================================================

def render_soft_constraints():
    """
    ソフト制約（S1学年集約・S2優先教科）の設定UIを表示する。
    render_generate()内の生成ボタン上部に配置する。
    """
    st.markdown("### 🎛️ ソフト制約設定")

    st.session_state["soft_grade_grouping"] = st.checkbox(
        "S1: 学年集約を有効にする",
        value=st.session_state.get("soft_grade_grouping", True),
        help=(
            "同一教員が同じ曜日に複数クラスを担当する場合、"
            "できるだけ同じ学年の授業が連続するように調整します"
        )
    )

    if st.session_state["soft_grade_grouping"]:
        st.session_state["soft_priority_subjects"] = st.multiselect(
            "S2: 優先教科（学年集約を特に強く適用する教科）",
            options=SUBJECTS,
            default=st.session_state.get("soft_priority_subjects", []),
            help=(
                "選択した教科は学年集約のペナルティが10倍になります。"
                "例: 理科・音楽・保体など準備が必要な教科を選択してください"
            )
        )

    st.markdown("---")
    st.markdown("#### 📚 教科別段階生成")
    st.session_state["soft_first_subjects"] = st.multiselect(
        "先に配置する教科（Phase 1）",
        options=SUBJECTS,
        default=st.session_state.get("soft_first_subjects", []),
        help=(
            "選択した教科を Phase 1 として先に全クラスに配置し、"
            "その後に残りの教科（Phase 2）を配置します。\n"
            "「📚 教科別段階生成」ボタンで実行してください。"
        )
    )


# ============================================================
# Part 3: バリデーション関数
# ============================================================

def validate_all_settings() -> list[str]:
    """
    生成前に全設定の整合性チェックを行い、警告メッセージのリストを返す。
    リストが空なら問題なし。
    """
    warnings_list = []
    classes        = get_all_classes()
    total_per_week = get_total_periods_per_week()

    # ① 教員未登録チェック
    if not st.session_state["teachers"]:
        warnings_list.append("教員が1人も登録されていません（STEP 2）")

    # ② 週コマ数の合計チェック
    for cls in classes:
        total = sum(st.session_state["weekly_periods"].get(cls, {}).values())
        if total != total_per_week:
            warnings_list.append(
                f"{cls}: 週コマ数の合計が {total} コマ "
                f"（週上限 {total_per_week} コマと不一致）"
            )

    # ③ 担当教員未設定チェック（週コマ数>0なのに担当なし）
    for cls in classes:
        for subj in SUBJECTS:
            weekly = st.session_state["weekly_periods"].get(cls, {}).get(subj, 0)
            if weekly > 0:
                teachers = st.session_state["assignments"].get(
                    cls, {}).get(subj, [])
                if not teachers:
                    warnings_list.append(
                        f"{cls} / {subj}: 週{weekly}コマ設定だが担当教員が未設定"
                    )

    # ④ 特別教室の教科名整合性チェック
    for room in st.session_state["special_rooms"]:
        if room.get("subject") not in SUBJECTS:
            warnings_list.append(
                f"特別教室「{room.get('name', '?')}」の対応教科が不正"
            )

    # ⑤ 少人数学級の同期グループ整合性チェック
    for sg_name, sg_data in st.session_state["small_group_classes"].items():
        for gi, group in enumerate(sg_data.get("sync_groups", [])):
            if len(group) < 2:
                warnings_list.append(
                    f"少人数学級「{sg_name}」のグループ{gi+1}: "
                    "2クラス以上必要です"
                )
            for cls in group:
                if cls not in sg_data.get("classes", []):
                    warnings_list.append(
                        f"少人数学級「{sg_name}」のグループ{gi+1}: "
                        f"「{cls}」が所属クラスに含まれていません"
                    )

    # ⑥ R7事前チェック（教員の週総授業コマ数が上限内か）
    for teacher in st.session_state["teachers"]:
        weekly_load = sum(
            st.session_state["weekly_periods"].get(cls, {}).get(subj, 0)
            for cls in classes
            for subj, tlist in st.session_state["assignments"].get(cls, {}).items()
            if teacher in tlist
        )
        if weekly_load == 0:
            continue
        max_weekly = sum(
            max(0, st.session_state["periods_per_day"].get(day, 6)
                - len(st.session_state["unavailable"].get(teacher, {}).get(day, []))
                - 1)
            for day in DAYS
        )
        if weekly_load > max_weekly:
            warnings_list.append(
                f"教員「{teacher}」: 週{weekly_load}コマ必要ですが"
                f"R7制約後の上限は{max_weekly}コマです（担当クラスを減らすか不在設定を見直してください）"
            )

    # ⑦ 手動コマのR1・R5チェック
    manual_tt = st.session_state.get("manual_timetable", {})
    if manual_tt:
        for day in DAYS:
            total_p = st.session_state["periods_per_day"].get(day, 6)
            for p in range(1, total_p + 1):
                seen_t: dict = {}
                for cls in classes:
                    subj = manual_tt.get(cls, {}).get(day, {}).get(p)
                    if not subj:
                        continue
                    # R5: 同日同教科
                    day_map = manual_tt.get(cls, {}).get(day, {})
                    same = [s for pk, s in day_map.items() if pk != p and s == subj]
                    if same:
                        warnings_list.append(
                            f"手動コマR5違反: {cls} {day}曜に「{subj}」が複数コマ設定されています")
                    # R1: 教員重複
                    for t in st.session_state["assignments"].get(cls, {}).get(subj, []):
                        if t in seen_t:
                            warnings_list.append(
                                f"手動コマR1違反: {t}先生が{day}曜{p}限に"
                                f"{seen_t[t]}と{cls}の両方に設定されています")
                        seen_t[t] = cls
                        # ⑧ 手動コマと不在コマの競合チェック（R4）
                        if p in st.session_state["unavailable"].get(t, {}).get(day, []):
                            warnings_list.append(
                                f"手動コマと不在コマが競合: {t}先生は{day}曜{p}限が不在設定ですが、"
                                f"{cls}の「{subj}」が手動配置されています（時間割を生成できません）"
                            )

    # ⑧ クラス間教科同期（R8）と担当教員（R1）の矛盾チェック
    for pair in st.session_state.get("class_subject_sync", []):
        c1, s1 = pair["class1"], pair["subject1"]
        c2, s2 = pair["class2"], pair["subject2"]
        teachers1 = set(st.session_state["assignments"].get(c1, {}).get(s1, []))
        teachers2 = set(st.session_state["assignments"].get(c2, {}).get(s2, []))
        conflict = teachers1 & teachers2
        if conflict:
            warnings_list.append(
                f"【R1×R8矛盾】{', '.join(conflict)} 先生が "
                f"「{c1} {s1}」と「{c2} {s2}」の両方に担当登録されています。"
                "同期制約により同じ時間に配置が必要ですが、教員重複禁止で不可能です。"
                f"STEP 5 で担当を修正してください（{c1} {s1}：{c1}側の先生のみ、"
                f"{c2} {s2}：{c2}側の先生のみ）。"
            )

    return warnings_list


# ============================================================
# Part 4: 制約チェック関数
# ============================================================

def is_valid_placement(
    timetable: dict,
    timetable_sg: dict,
    cls: str,
    day: str,
    period: int,
    subject: str,
    assignments: dict,
    special_rooms: list,
    unavailable: dict,
    periods_per_day: dict,
    small_group_classes: dict,
    weekly_remaining: dict
) -> bool:
    """R1〜R7を全チェックする。すべてパスしたらTrue。"""

    teachers = assignments.get(cls, {}).get(subject, [])

    # ── R1: 教員重複禁止 ─────────────────────────────────────
    for t in teachers:
        for other_cls, schedule in timetable.items():
            if other_cls == cls:
                continue
            other_subj = schedule.get(day, {}).get(period)
            if other_subj:
                if t in assignments.get(other_cls, {}).get(other_subj, []):
                    return False
        for sg_name, sg_schedule in timetable_sg.items():
            sg_subj = sg_schedule.get(day, {}).get(period)
            if sg_subj:
                sg_teachers = small_group_classes.get(sg_name, {}).get("teachers", [])
                if t in sg_teachers:
                    return False

    # ── R3: 特別教室同時使用制限 ─────────────────────────────
    for room in special_rooms:
        if room["subject"] != subject:
            continue
        count = sum(
            1 for other_cls, schedule in timetable.items()
            if other_cls != cls and schedule.get(day, {}).get(period) == subject
        )
        if count >= room["capacity"]:
            return False

    # ── R4: 教員不在コマ尊重 ─────────────────────────────────
    for t in teachers:
        if period in unavailable.get(t, {}).get(day, []):
            return False

    # ── R5: 同一クラス・同一曜日に同一教科は1コマまで ───────
    for p, s in timetable.get(cls, {}).get(day, {}).items():
        if p != period and s == subject:
            return False

    # ── R6: 少人数学級同期制約 ───────────────────────────────
    for sg_name, sg_data in small_group_classes.items():
        groups = sg_data.get("sync_groups") or [sg_data.get("classes", [])]
        for group in groups:
            if cls not in group:
                continue
            other_subjects = [
                timetable.get(oc, {}).get(day, {}).get(period)
                for oc in group if oc != cls
            ]
            other_subjects = [s for s in other_subjects if s]
            if not other_subjects:
                continue
            this_is_common = subject in COMMON_SUBJECTS
            for os_ in other_subjects:
                if (os_ in COMMON_SUBJECTS) != this_is_common:
                    return False

    # ── R7: 各教員の当日最低1コマ空き ───────────────────────
    total_periods_today = periods_per_day.get(day, 6)
    for t in teachers:
        assigned_today = sum(
            1 for oc, sc in timetable.items()
            for p_ in range(1, total_periods_today + 1)
            if sc.get(day, {}).get(p_) and
               t in assignments.get(oc, {}).get(sc[day].get(p_, ""), [])
        )
        assigned_today += sum(
            1 for sg_name, sg_sc in timetable_sg.items()
            for p_ in range(1, total_periods_today + 1)
            if sg_sc.get(day, {}).get(p_) and
               t in small_group_classes.get(sg_name, {}).get("teachers", [])
        )
        unavail_today = len(unavailable.get(t, {}).get(day, []))
        max_assignable = total_periods_today - unavail_today - 1
        if assigned_today >= max_assignable:
            return False

    return True


# ============================================================
# Part 4: 前処理関数
# ============================================================

def preprocess(
    timetable: dict,
    timetable_sg: dict,
    manual_timetable: dict,
    manual_timetable_sg: dict,
    weekly_periods: dict,
    weekly_periods_sg: dict
) -> tuple[dict, dict]:
    """
    手動入力コマをロックし、残り週コマ数を計算して返す。
    戻り値: (weekly_remaining, weekly_remaining_sg)
    """
    # 通常学級の残り週コマ数を初期化
    weekly_remaining = {
        cls: dict(weekly_periods.get(cls, {}))
        for cls in timetable
    }

    # 手動入力コマをロック
    for cls, day_map in manual_timetable.items():
        if cls not in timetable:
            continue
        for day, period_map in day_map.items():
            for period, subj in period_map.items():
                timetable[cls].setdefault(day, {})[period] = subj
                if subj in weekly_remaining.get(cls, {}):
                    weekly_remaining[cls][subj] = max(
                        0, weekly_remaining[cls][subj] - 1
                    )

    # 少人数学級の残り週コマ数を初期化
    weekly_remaining_sg = {
        sg: dict(weekly_periods_sg.get(sg, {}))
        for sg in timetable_sg
    }

    # 少人数手動入力コマをロック
    for sg_name, day_map in manual_timetable_sg.items():
        if sg_name not in timetable_sg:
            continue
        for day, period_map in day_map.items():
            for period, subj in period_map.items():
                timetable_sg[sg_name].setdefault(day, {})[period] = subj
                if subj in weekly_remaining_sg.get(sg_name, {}):
                    weekly_remaining_sg[sg_name][subj] = max(
                        0, weekly_remaining_sg[sg_name][subj] - 1
                    )

    return weekly_remaining, weekly_remaining_sg


# ============================================================
# Part 4: OR-Tools フォールバックソルバー
# ============================================================

def solve_ortools(
    timetable: dict,
    timetable_sg: dict,
    weekly_remaining: dict,
    assignments: dict,
    special_rooms: list,
    unavailable: dict,
    periods_per_day: dict,
    small_group_classes: dict,
    priority_subjects: list,
    random_seed: int = 0,
    time_limit: float = 60.0
) -> str:
    """
    Google OR-Tools CP-SAT で時間割を生成する。
    time_limit: ソルバーの最大実行秒数（デフォルト60秒）
    戻り値: "solved" / "failed"
    """
    try:
        from ortools.sat.python import cp_model

        model   = cp_model.CpModel()
        classes = list(timetable.keys())
        x       = {}  # x[cls][day][p][subj] ∈ {0,1}

        # ── 変数定義 ─────────────────────────────────────────
        for cls in classes:
            x[cls] = {}
            for day in DAYS:
                x[cls][day] = {}
                for p in range(1, periods_per_day.get(day, 6) + 1):
                    x[cls][day][p] = {
                        subj: model.NewBoolVar(f"x_{cls}_{day}_{p}_{subj}")
                        for subj in SUBJECTS
                    }

        # ── 手動ロック済みコマを固定 ─────────────────────────
        for cls in classes:
            for day in DAYS:
                for p in range(1, periods_per_day.get(day, 6) + 1):
                    locked = timetable[cls].get(day, {}).get(p)
                    if locked:
                        model.Add(x[cls][day][p][locked] == 1)
                        for subj in SUBJECTS:
                            if subj != locked:
                                model.Add(x[cls][day][p][subj] == 0)

        # ── 基本制約: 各コマに高々1教科 ─────────────────────
        for cls in classes:
            for day in DAYS:
                for p in range(1, periods_per_day.get(day, 6) + 1):
                    model.AddAtMostOne(x[cls][day][p][s] for s in SUBJECTS)

        # ── R1: 教員重複禁止 ─────────────────────────────────
        for t in st.session_state["teachers"]:
            for day in DAYS:
                for p in range(1, periods_per_day.get(day, 6) + 1):
                    tvars = [
                        x[cls][day][p][subj]
                        for cls in classes
                        for subj in SUBJECTS
                        if t in assignments.get(cls, {}).get(subj, [])
                    ]
                    if tvars:
                        model.AddAtMostOne(tvars)

        # ── R2: 週コマ数完全消化 ─────────────────────────────
        for cls in classes:
            for subj in SUBJECTS:
                target = weekly_remaining.get(cls, {}).get(subj, 0)
                pvars = [
                    x[cls][day][p][subj]
                    for day in DAYS
                    for p in range(1, periods_per_day.get(day, 6) + 1)
                    if not timetable[cls].get(day, {}).get(p)
                ]
                if pvars:
                    model.Add(sum(pvars) == target)

        # ── R3: 特別教室同時使用制限 ─────────────────────────
        for room in special_rooms:
            subj = room["subject"]
            cap  = room["capacity"]
            for day in DAYS:
                for p in range(1, periods_per_day.get(day, 6) + 1):
                    model.Add(sum(x[cls][day][p][subj] for cls in classes) <= cap)

        # ── R4: 教員不在コマ ─────────────────────────────────
        for t in st.session_state["teachers"]:
            for day in DAYS:
                for p in unavailable.get(t, {}).get(day, []):
                    if p > periods_per_day.get(day, 6):
                        continue
                    for cls in classes:
                        for subj in SUBJECTS:
                            if t in assignments.get(cls, {}).get(subj, []):
                                model.Add(x[cls][day][p][subj] == 0)

        # ── R5: 同曜日・同教科1コマまで ─────────────────────
        for cls in classes:
            for day in DAYS:
                for subj in SUBJECTS:
                    model.Add(sum(
                        x[cls][day][p][subj]
                        for p in range(1, periods_per_day.get(day, 6) + 1)
                    ) <= 1)

        # ── R8: クラス間教科同期（ハード制約・双方向） ────────
        # x[c1][d][p][s1] == x[c2][d][p][s2] を全スロットに課す
        for pair in st.session_state.get("class_subject_sync", []):
            c1, s1 = pair["class1"], pair["subject1"]
            c2, s2 = pair["class2"], pair["subject2"]
            if c1 not in classes or c2 not in classes:
                continue
            for day in DAYS:
                for p in range(1, periods_per_day.get(day, 6) + 1):
                    model.Add(x[c1][day][p][s1] == x[c2][day][p][s2])

        # ── 目的関数ペナルティ項（R6ソフト＋S1学年集約） ────────
        penalty_terms = []

        # ── R6: 少人数学級同期（ソフト制約・ペナルティ） ────────
        # ハード制約から外し、違反1件あたり大きなペナルティを課す。
        # これにより「できるだけ揃える」方向に最適化しつつ、
        # 揃えられない場合も時間割全体は生成できる。
        R6_PENALTY = 1000
        for sg_name, sg_data in small_group_classes.items():
            groups = sg_data.get("sync_groups") or [sg_data.get("classes", [])]
            for group in groups:
                valid_group = [c for c in group if c in classes]
                if len(valid_group) < 2:
                    continue
                for day in DAYS:
                    for p in range(1, periods_per_day.get(day, 6) + 1):
                        # is_common_vars[c] = 1 ならそのクラスはこの時限に共通教科
                        is_common_vars = {}
                        for c in valid_group:
                            iv = model.NewBoolVar(f"ic_{sg_name}_{c}_{day}_{p}")
                            cs = sum(x[c][day][p][s] for s in COMMON_SUBJECTS)
                            model.Add(cs >= 1).OnlyEnforceIf(iv)
                            model.Add(cs == 0).OnlyEnforceIf(iv.Not())
                            is_common_vars[c] = iv
                        # 隣接するクラスペアで不一致のときペナルティ変数を立てる
                        # ── R6ペナルティ変数: mismatch = ref XOR c ───────
                        # AddBoolXOr([a,b,c]) は a XOR b XOR c = 1 を意味する
                        # mismatch = ref XOR c を表すには
                        # ref XOR c XOR mismatch.Not() = 1 とする
                        ref = valid_group[0]
                        for c in valid_group[1:]:
                            mismatch = model.NewBoolVar(
                                f"r6mismatch_{sg_name}_{ref}_{c}_{day}_{p}")
                            model.AddBoolXOr(
                                [is_common_vars[ref],
                                 is_common_vars[c],
                                 mismatch.Not()]
                            )
                            penalty_terms.append(R6_PENALTY * mismatch)

        # ── R7: 教員1日最低1コマ空き ─────────────────────────
        for t in st.session_state["teachers"]:
            for day in DAYS:
                total_today = periods_per_day.get(day, 6)
                unavail_cnt = len(unavailable.get(t, {}).get(day, []))
                available   = total_today - unavail_cnt   # 出勤可能コマ数
                # 少人数学級の手動ロック済みコマ数を考慮して上限を減算
                sg_cnt = sum(
                    1 for sg_name, sg_sched in timetable_sg.items()
                    for p in range(1, total_today + 1)
                    if sg_sched.get(day, {}).get(p)
                    and t in small_group_classes.get(sg_name, {}).get("teachers", [])
                    and p not in unavailable.get(t, {}).get(day, [])
                )
                max_ok      = available - 1 - sg_cnt     # 最低1コマ空き（少人数分を減算）
                if max_ok <= 0:
                    continue  # 出勤不可またはすでに上限の曜日はR7対象外
                tvars = [
                    x[cls][day][p][subj]
                    for cls in classes
                    for p in range(1, total_today + 1)
                    for subj in SUBJECTS
                    if t in assignments.get(cls, {}).get(subj, [])
                    and p not in unavailable.get(t, {}).get(day, [])
                ]
                if tvars:
                    model.Add(sum(tvars) <= max_ok)

        # ── S1: 学年集約ペナルティ（時限集約 + 曜日集約） ──────────
        if st.session_state.get("soft_grade_grouping", True):
            PW, NW = 200, 20

            # S1b: 同学年・同教科は同じ曜日に集約するペナルティ
            # has_on_day[c][s][d] = クラスcが教科sを曜日dに持つか
            has_on_day = {}
            for c in classes:
                has_on_day[c] = {}
                for s in SUBJECTS:
                    has_on_day[c][s] = {}
                    for d in DAYS:
                        dp = periods_per_day.get(d, 6)
                        xvars = [x[c][d][p][s] for p in range(1, dp + 1)]
                        hv = model.NewBoolVar(f"hod_{c}_{s}_{d}")
                        model.AddBoolOr(xvars).OnlyEnforceIf(hv)
                        model.AddBoolAnd([xv.Not() for xv in xvars]).OnlyEnforceIf(hv.Not())
                        has_on_day[c][s][d] = hv

            # 同学年クラスのペアで、同教科が異なる曜日に配置されたらペナルティ
            grade_cls_map: dict[int, list] = {}
            for c in classes:
                grade_cls_map.setdefault(get_grade(c), []).append(c)

            for gcls in grade_cls_map.values():
                for i, c1 in enumerate(gcls):
                    for c2 in gcls[i + 1:]:
                        for s in SUBJECTS:
                            w = PW if s in priority_subjects else NW
                            for d in DAYS:
                                # c1とc2でsの曜日配置が食い違うときペナルティ
                                mismatch = model.NewBoolVar(
                                    f"daymm_{c1}_{c2}_{s}_{d}")
                                model.AddBoolXOr([
                                    has_on_day[c1][s][d],
                                    has_on_day[c2][s][d],
                                    mismatch.Not()
                                ])
                                penalty_terms.append(w * mismatch)

            # S3: 学年交互防止（全時限で同一教員が連続して異学年を担当しない）
            # 優先教科が絡む遷移は S3_PW、それ以外は S3_NW でペナルティを課す
            S3_PW, S3_NW = 200, 20
            S3_TARGET = list(range(1, max(periods_per_day.values(), default=8) + 1))
            grades_in_school = sorted({get_grade(c) for c in classes})

            if len(grades_in_school) >= 2 and len(S3_TARGET) >= 2:
                # teaches_grade_var[t][d][p][g]: 全教科版
                teaches_grade_var = {}
                # teaches_prio_grade[t][d][p][g]: 優先教科のみ版
                teaches_prio_grade = {}
                for t in st.session_state["teachers"]:
                    teaches_grade_var[t] = {}
                    teaches_prio_grade[t] = {}
                    for d in DAYS:
                        teaches_grade_var[t][d] = {}
                        teaches_prio_grade[t][d] = {}
                        for p in S3_TARGET:
                            teaches_grade_var[t][d][p] = {}
                            teaches_prio_grade[t][d][p] = {}
                            for g in grades_in_school:
                                # 全教科
                                tvars_g = [
                                    x[c][d][p][s]
                                    for c in classes if get_grade(c) == g
                                    for s in SUBJECTS
                                    if t in assignments.get(c, {}).get(s, [])
                                    and p <= periods_per_day.get(d, 6)
                                ]
                                tg = model.NewBoolVar(f"tg_{t}_{d}_{p}_{g}")
                                if tvars_g:
                                    model.AddBoolOr(tvars_g).OnlyEnforceIf(tg)
                                    model.AddBoolAnd(
                                        [v.Not() for v in tvars_g]
                                    ).OnlyEnforceIf(tg.Not())
                                else:
                                    model.Add(tg == 0)
                                teaches_grade_var[t][d][p][g] = tg

                                # 優先教科のみ
                                pvars_g = [
                                    x[c][d][p][s]
                                    for c in classes if get_grade(c) == g
                                    for s in priority_subjects
                                    if t in assignments.get(c, {}).get(s, [])
                                    and p <= periods_per_day.get(d, 6)
                                ]
                                tpg = model.NewBoolVar(f"tpg_{t}_{d}_{p}_{g}")
                                if pvars_g:
                                    model.AddBoolOr(pvars_g).OnlyEnforceIf(tpg)
                                    model.AddBoolAnd(
                                        [v.Not() for v in pvars_g]
                                    ).OnlyEnforceIf(tpg.Not())
                                else:
                                    model.Add(tpg == 0)
                                teaches_prio_grade[t][d][p][g] = tpg

                # 連続する時限 (p, p+1) で学年が切り替わるとペナルティ
                # 優先教科が絡む遷移は S3_PW、それ以外は S3_NW
                for t in st.session_state["teachers"]:
                    for d in DAYS:
                        for p in S3_TARGET[:-1]:   # (1,2),(2,3),(3,4)
                            p2 = p + 1
                            if p2 not in S3_TARGET:
                                continue
                            for g1 in grades_in_school:
                                for g2 in grades_in_school:
                                    if g1 == g2:
                                        continue
                                    tg1 = teaches_grade_var[t][d][p][g1]
                                    tg2 = teaches_grade_var[t][d][p2][g2]
                                    # 学年遷移変数
                                    pv = model.NewBoolVar(
                                        f"s3_{t}_{d}_{p}_{g1}_{g2}")
                                    model.AddBoolAnd([tg1, tg2]).OnlyEnforceIf(pv)
                                    model.AddBoolOr(
                                        [tg1.Not(), tg2.Not()]
                                    ).OnlyEnforceIf(pv.Not())
                                    # 基本ペナルティ（全遷移に S3_NW）
                                    penalty_terms.append(S3_NW * pv)
                                    # 優先教科が絡む場合は追加ペナルティ
                                    if priority_subjects:
                                        tpg1 = teaches_prio_grade[t][d][p][g1]
                                        pv_p = model.NewBoolVar(
                                            f"s3p_{t}_{d}_{p}_{g1}_{g2}")
                                        model.AddBoolAnd(
                                            [pv, tpg1]
                                        ).OnlyEnforceIf(pv_p)
                                        model.AddBoolOr(
                                            [pv.Not(), tpg1.Not()]
                                        ).OnlyEnforceIf(pv_p.Not())
                                        # 追加分: PW - NW = 9
                                        penalty_terms.append(
                                            (S3_PW - S3_NW) * pv_p)

        if penalty_terms:
            model.Minimize(sum(penalty_terms))

        # ── ソルバー実行 ─────────────────────────────────────
        solver = cp_model.CpSolver()
        solver.parameters.max_time_in_seconds  = float(time_limit)
        solver.parameters.num_search_workers   = 4
        if random_seed != 0:
            solver.parameters.random_seed = random_seed % (2**31 - 1)
        status = solver.Solve(model)

        if status in (cp_model.OPTIMAL, cp_model.FEASIBLE):
            for cls in classes:
                for day in DAYS:
                    for p in range(1, periods_per_day.get(day, 6) + 1):
                        if timetable[cls].get(day, {}).get(p):
                            continue
                        for subj in SUBJECTS:
                            if solver.Value(x[cls][day][p][subj]) == 1:
                                timetable[cls].setdefault(day, {})[p] = subj
                                # weekly_remaining を減算して整合性を保つ
                                if subj in weekly_remaining.get(cls, {}):
                                    weekly_remaining[cls][subj] = max(
                                        0, weekly_remaining[cls][subj] - 1)
                                break
            return "solved"
        return "failed"

    except Exception as e:
        st.error(f"OR-Toolsエラー: {e}")
        return "failed"



# ============================================================
# Part 4: 学年別段階生成
# ============================================================

def solve_grade_by_grade(
    timetable: dict,
    timetable_sg: dict,
    weekly_remaining: dict,
    assignments: dict,
    special_rooms: list,
    unavailable: dict,
    periods_per_day: dict,
    small_group_classes: dict,
    priority_subjects: list,
    ort_time_limit: float = 60.0
) -> str:
    """
    学年ごとに順番に解く段階的生成。

    クロス学年の制約処理:
    ・R1（教員重複）: 前学年の使用済みコマを extra_unavail として次学年に引き継ぐ
    ・R3（教室容量）: 前学年の per-slot 使用数を特別教室の容量から引いて渡す
    ・少人数同期のある学年を先に解く（制約が厳しいため）
    戻り値: "solved" / "failed_grade{学年番号}"
    """
    from collections import defaultdict

    # ── 学年別クラスグループ ─────────────────────────────────
    grade_groups: dict[int, list] = defaultdict(list)
    for cls in timetable:
        grade_groups[get_grade(cls)].append(cls)

    # ── 少人数同期グループを学年別にフィルタリング ─────────────
    def sg_for_classes(cls_list):
        result = {}
        for sg_name, sg_data in small_group_classes.items():
            in_list = [c for c in sg_data.get("classes", []) if c in cls_list]
            if not in_list:
                continue
            filtered_groups = [
                [c for c in g if c in cls_list]
                for g in (sg_data.get("sync_groups") or [sg_data.get("classes", [])])
                if sum(1 for c in g if c in cls_list) >= 2
            ]
            if filtered_groups:
                result[sg_name] = {**sg_data,
                                   "classes": in_list,
                                   "sync_groups": filtered_groups}
        return result

    # ── 処理順: 少人数同期のある学年を優先 ──────────────────
    sorted_grades = sorted(grade_groups.keys(),
                           key=lambda g: -len(sg_for_classes(grade_groups[g])))

    # ── 追加不在（前学年で使用した教員スロット）───────────────
    # teacher -> day -> [period, ...]
    extra_unavail: dict[str, dict[str, list]] = defaultdict(
        lambda: defaultdict(list))

    # ── 教室スロット使用数（前学年で埋まった分）────────────────
    room_slot_used: dict[str, dict[str, dict[int, int]]] = {
        room["subject"]: defaultdict(lambda: defaultdict(int))
        for room in special_rooms
    }

    for grade in sorted_grades:
        grade_classes = grade_groups[grade]

        # この学年のtimetable / weekly_remaining を切り出す
        tt_g  = {cls: timetable[cls]         for cls in grade_classes}
        wr_g  = {cls: weekly_remaining[cls]  for cls in grade_classes}
        sg_g  = sg_for_classes(grade_classes)

        # 前学年の使用済みスロットを不在として合成
        merged_unavail = {}
        all_teachers_in_school = set(
            t for cls_asgn in assignments.values()
            for subj_teachers in cls_asgn.values()
            for t in subj_teachers
        )
        for t in all_teachers_in_school:
            base = dict(unavailable.get(t, {}))
            for day_, periods_ in extra_unavail[t].items():
                existing = list(base.get(day_, []))
                for p in periods_:
                    if p not in existing:
                        existing.append(p)
                base[day_] = existing
            merged_unavail[t] = base

        # 特別教室の実効容量（前学年分を引く）
        adj_rooms = []
        for room in special_rooms:
            subj = room["subject"]
            adj_rooms.append(room)  # capacity は OR-Tools 側で変数制約として処理

        wr_g2 = {cls: dict(weekly_remaining[cls]) for cls in grade_classes}
        # 手動コマ再ロック
        for cls in grade_classes:
            for day_, period_map in timetable[cls].items():
                for p_, s_ in period_map.items():
                    if s_:
                        tt_g[cls][day_][p_] = s_
                        if s_ in wr_g2[cls]:
                            wr_g2[cls][s_] = max(0, wr_g2[cls][s_] - 1)
        result = solve_ortools(
            tt_g, timetable_sg,
            wr_g2, assignments,
            special_rooms, merged_unavail,
            periods_per_day, sg_g,
            priority_subjects, time_limit=ort_time_limit
        )
        if result == "solved":
            for cls in grade_classes:
                weekly_remaining[cls] = wr_g2[cls]

        if result != "solved":
            return f"failed_grade{grade}"

        # 解をメインtimetableに反映
        for cls in grade_classes:
            timetable[cls] = tt_g[cls]
            weekly_remaining[cls] = wr_g[cls]

        # ── 教員使用スロットと教室使用スロットを次学年へ引き継ぐ ──
        for cls in grade_classes:
            for day in DAYS:
                for p in range(1, periods_per_day.get(day, 6) + 1):
                    subj = timetable[cls][day].get(p)
                    if not subj:
                        continue
                    # R1: 教員スロット蓄積
                    for t in assignments.get(cls, {}).get(subj, []):
                        if p not in extra_unavail[t][day]:
                            extra_unavail[t][day].append(p)
                    # R3: 教室使用カウント
                    for room in special_rooms:
                        if room["subject"] == subj:
                            room_slot_used[subj][day][p] += 1

    return "solved"


# ============================================================
# Part 4: 教科別段階生成
# ============================================================

def solve_subject_by_subject(
    timetable: dict,
    timetable_sg: dict,
    weekly_remaining: dict,
    assignments: dict,
    special_rooms: list,
    unavailable: dict,
    periods_per_day: dict,
    small_group_classes: dict,
    priority_subjects: list,
    first_subjects: list,
    ort_time_limit: float = 60.0
) -> str:
    """
    選択した教科を先に配置し、残りの教科を後から配置する段階的生成。

    Phase 1: first_subjects を全クラスに配置
    Phase 2: 残りの教科を配置（Phase 1 の教員スロットを引き継ぐ）

    戻り値: "solved" / "failed_phase1" / "failed_phase2"
    """
    remaining_subjects = [s for s in SUBJECTS if s not in first_subjects]

    for phase_idx, phase_subjects in enumerate([first_subjects, remaining_subjects], 1):
        if not phase_subjects:
            continue

        # この段階で配置する教科のみに絞った weekly_remaining
        wr_phase = {
            cls: {s: cnt for s, cnt in weekly_remaining[cls].items()
                  if s in phase_subjects}
            for cls in weekly_remaining
        }

        # 全クラスで残コマが 0 なら skip
        if all(v == 0 for rem in wr_phase.values() for v in rem.values()):
            continue

        result = solve_ortools(
            timetable, timetable_sg,
            wr_phase, assignments,
            special_rooms, unavailable,
            periods_per_day, small_group_classes,
            priority_subjects, time_limit=ort_time_limit
        )

        if result != "solved":
            return f"failed_phase{phase_idx}"

        # timetable は in-place 更新済みなので次フェーズでそのまま使う

    return "solved"


# ============================================================
# Part 4: エラーレポート
# ============================================================

def report_r6_violations(timetable: dict):
    """
    生成済み時間割のR6ソフト制約の達成状況を集計して表示する。
    違反 = 同期グループ内で、ある時限に共通教科と非共通教科が混在しているコマ。
    """
    small_group_classes = st.session_state.get("small_group_classes", {})
    if not small_group_classes:
        return

    total_slots   = 0  # チェック対象スロット（ペア×時限）数
    violations    = []  # (sg_name, day, period, {cls: subj}) のリスト

    for sg_name, sg_data in small_group_classes.items():
        groups = sg_data.get("sync_groups") or [sg_data.get("classes", [])]
        for group in groups:
            valid_group = [c for c in group if c in timetable]
            if len(valid_group) < 2:
                continue
            periods_per_day = st.session_state["periods_per_day"]
            for day in DAYS:
                for p in range(1, periods_per_day.get(day, 6) + 1):
                    subjects = {
                        c: timetable[c][day].get(p)
                        for c in valid_group
                    }
                    filled = {c: s for c, s in subjects.items() if s}
                    if len(filled) < 2:
                        continue
                    total_slots += 1
                    flags = [s in COMMON_SUBJECTS for s in filled.values()]
                    if len(set(flags)) > 1:  # 混在
                        violations.append((sg_name, day, p, subjects))

    if total_slots == 0:
        return

    ok_count = total_slots - len(violations)
    rate = ok_count / total_slots * 100

    if not violations:
        st.success(
            f"✅ R6（少人数同期）: 全 {total_slots} スロット一致 "
            f"（達成率 100%）"
        )
    else:
        st.warning(
            f"⚠️ R6（少人数同期）: {ok_count}/{total_slots} スロット一致 "
            f"（達成率 {rate:.0f}%）　違反 {len(violations)} 件"
        )
        with st.expander("R6違反の詳細を見る", expanded=False):
            for sg_name, day, p, subjects in violations:
                detail = "　".join(
                    f"{c}={'★' if s in COMMON_SUBJECTS else ''}{s}"
                    for c, s in subjects.items() if s
                )
                st.markdown(f"- **{sg_name}** {day}{p}限：{detail}")


def report_conflict(timetable: dict, weekly_remaining: dict):
    """解が見つからなかった場合に未配置コマと改善策を表示する。"""
    st.error("❌ 時間割を完成させることができませんでした")

    unplaced = [
        f"・{cls} ／ {subj}: あと **{rem}** コマ未配置"
        for cls, subj_map in weekly_remaining.items()
        for subj, rem in subj_map.items()
        if rem > 0
    ]
    if unplaced:
        st.markdown("**未配置コマ一覧:**")
        for item in unplaced:
            st.markdown(item)

    st.info(
        "💡 **改善策の例**\n"
        "- 週コマ数が多い教科を減らす（STEP 4）\n"
        "- 教員の不在コマを見直す（STEP 6）\n"
        "- 特別教室の同時使用可能クラス数を増やす（STEP 7）\n"
        "- 少人数学級の同期グループを再設定する（STEP 9）\n"
        "- 手動入力コマを減らす（STEP 8）"
    )


# ============================================================
# Part 4: 時間割表示
# ============================================================

def display_class_timetable(
    cls: str,
    timetable: dict,
    manual_timetable: dict,
    is_small_group: bool = False
):
    """クラスの時間割をグリッド表示する。手動=水色・自動=薄緑。"""
    max_p = get_max_periods()
    data  = {}

    for day in DAYS:
        day_periods = st.session_state["periods_per_day"].get(day, 6)
        col_data = []
        for p in range(1, max_p + 1):
            if p > day_periods:
                col_data.append("─")
                continue
            subj = timetable.get(cls, {}).get(day, {}).get(p, "")
            if not subj:
                col_data.append("")
                continue
            if is_small_group:
                teachers = st.session_state["small_group_classes"] \
                    .get(cls, {}).get("teachers", [])
            else:
                teachers = get_teachers_for_slot(cls, subj)
            cell = f"{subj}\n{'・'.join(teachers)}" if teachers else subj
            col_data.append(cell)
        data[day] = col_data

    df = pd.DataFrame(data, index=[f"{p}限" for p in range(1, max_p + 1)])
    st.dataframe(df, use_container_width=True)
    st.caption("🔵 水色=手動入力　🟢 薄緑=自動生成")


def display_teacher_timetable(
    teacher: str,
    timetable: dict,
    timetable_sg: dict
):
    """教員の時間割をグリッド表示する。空き・不在も色分け。"""
    max_p = get_max_periods()
    data  = {}

    for day in DAYS:
        day_periods = st.session_state["periods_per_day"].get(day, 6)
        col_data = []
        for p in range(1, max_p + 1):
            if p > day_periods:
                col_data.append("─")
                continue
            if p in st.session_state["unavailable"].get(teacher, {}).get(day, []):
                col_data.append("🚫不在")
                continue
            cell = "空き"
            for cls, schedule in timetable.items():
                subj = schedule.get(day, {}).get(p)
                if subj and teacher in get_teachers_for_slot(cls, subj):
                    cell = f"{cls}\n{subj}"
                    break
            if cell == "空き":
                for sg_name, sg_schedule in timetable_sg.items():
                    subj = sg_schedule.get(day, {}).get(p)
                    if subj and teacher in st.session_state["small_group_classes"] \
                            .get(sg_name, {}).get("teachers", []):
                        cell = f"{sg_name}\n{subj}"
                        break
            col_data.append(cell)
        data[day] = col_data

    df = pd.DataFrame(data, index=[f"{p}限" for p in range(1, max_p + 1)])
    st.dataframe(df, use_container_width=True)
    st.caption("🚫 不在　空欄=空きコマ")


def display_all_teachers_timetable(timetable: dict, timetable_sg: dict):
    """全教員の時間割を一覧表示する。縦=教員、横=曜日×時限。"""
    teachers      = st.session_state["teachers"]
    periods_per_day = st.session_state["periods_per_day"]
    unavailable   = st.session_state["unavailable"]
    small_group_classes = st.session_state["small_group_classes"]

    # 列ラベルを生成（例: 月1, 月2, ..., 金6）
    columns = [
        f"{day}{p}"
        for day in DAYS
        for p in range(1, periods_per_day.get(day, 6) + 1)
    ]

    rows = {}
    for teacher in teachers:
        row = []
        for day in DAYS:
            day_periods = periods_per_day.get(day, 6)
            for p in range(1, day_periods + 1):
                if p in unavailable.get(teacher, {}).get(day, []):
                    row.append("🚫不在")
                    continue
                cell = ""
                for cls, schedule in timetable.items():
                    subj = schedule.get(day, {}).get(p)
                    if subj and teacher in get_teachers_for_slot(cls, subj):
                        cell = f"{cls} {subj}"
                        break
                if not cell:
                    for sg_name, sg_schedule in timetable_sg.items():
                        subj = sg_schedule.get(day, {}).get(p)
                        if subj and teacher in small_group_classes.get(
                                sg_name, {}).get("teachers", []):
                            cell = f"{sg_name} {subj}"
                            break
                row.append(cell)
        rows[teacher] = row

    df = pd.DataFrame(rows, index=columns).T
    st.dataframe(df, use_container_width=True)
    st.caption("🚫 不在　空欄=空きコマ")


# ============================================================
# Part 5: 手動編集 ── コマ交換機能
# ============================================================

def is_valid_swap(
    timetable: dict,
    timetable_sg: dict,
    cls1: str, day1: str, p1: int,
    cls2: str, day2: str, p2: int
) -> bool:
    """
    (cls1, day1, p1) の教科と (cls2, day2, p2) の教科を交換したとき、
    全制約（R1/R3/R4/R5/R7/R8）を満たすかチェックする。
    """
    if cls1 == cls2 and day1 == day2 and p1 == p2:
        return False

    periods_per_day = st.session_state["periods_per_day"]
    unavailable     = st.session_state["unavailable"]
    special_rooms   = st.session_state["special_rooms"]
    sync_pairs      = st.session_state.get("class_subject_sync", [])
    small_group_cls = st.session_state["small_group_classes"]

    subj1 = timetable.get(cls1, {}).get(day1, {}).get(p1)  # 移動元教科（必須）
    subj2 = timetable.get(cls2, {}).get(day2, {}).get(p2)  # 移動先教科（空きも可）

    if not subj1:
        return False
    if cls1 == cls2 and subj1 == subj2:
        return False
    if p1 > periods_per_day.get(day1, 6) or p2 > periods_per_day.get(day2, 6):
        return False

    same_cls_same_day = (cls1 == cls2 and day1 == day2)

    # ── R4: 教員不在 ──────────────────────────────────────────
    if subj2:
        for t in get_teachers_for_slot(cls1, subj2):
            if p1 in unavailable.get(t, {}).get(day1, []):
                return False
    for t in get_teachers_for_slot(cls2, subj1):
        if p2 in unavailable.get(t, {}).get(day2, []):
            return False

    # ── R5: 同日同教科（クラス内）────────────────────────────
    if subj2:
        for p_x, s_x in timetable.get(cls1, {}).get(day1, {}).items():
            if p_x == p1:
                continue
            if same_cls_same_day and p_x == p2:
                continue  # p2 は swap 後 subj1 になる
            if s_x == subj2:
                return False
    for p_x, s_x in timetable.get(cls2, {}).get(day2, {}).items():
        if p_x == p2:
            continue
        if same_cls_same_day and p_x == p1:
            continue  # p1 は swap 後 subj2 になる
        if s_x == subj1:
            return False

    # ── R1: 教員重複 ─────────────────────────────────────────
    def busy_at(day, p, exclude_cls):
        busy = set()
        for cls, schedule in timetable.items():
            if cls == exclude_cls:
                continue
            s = schedule.get(day, {}).get(p)
            if s:
                for t in get_teachers_for_slot(cls, s):
                    busy.add(t)
        for sg_name, schedule in timetable_sg.items():
            s = schedule.get(day, {}).get(p)
            if s:
                for t in small_group_cls.get(sg_name, {}).get("teachers", []):
                    busy.add(t)
        return busy

    busy_src = busy_at(day1, p1, cls1)
    busy_dst = busy_at(day2, p2, cls2)

    t1_set = set(get_teachers_for_slot(cls2, subj1))
    t2_set = set(get_teachers_for_slot(cls1, subj2)) if subj2 else set()

    if t2_set & busy_src:
        return False
    if t1_set & busy_dst:
        return False
    # 同一時限 swap: 2つの新配置が教員を共有してはいけない
    if day1 == day2 and p1 == p2 and t2_set & t1_set:
        return False

    # ── R3: 特別教室容量 ──────────────────────────────────────
    for room in special_rooms:
        subj_r = room["subject"]
        cap    = room["capacity"]
        if subj2 == subj_r:
            count = sum(
                1 for cls, sched in timetable.items()
                if cls != cls1
                and sched.get(day1, {}).get(p1) == subj_r
                and not (cls == cls2 and day1 == day2 and p1 == p2)
            )
            if count + 1 > cap:
                return False
        if subj1 == subj_r:
            count = sum(
                1 for cls, sched in timetable.items()
                if cls != cls2
                and sched.get(day2, {}).get(p2) == subj_r
                and not (cls == cls1 and day1 == day2 and p1 == p2)
            )
            if count + 1 > cap:
                return False

    # ── R7: 教員日内コマ数上限 ───────────────────────────────
    teacher_day_cnt: dict = {}
    for cls, schedule in timetable.items():
        for day, periods in schedule.items():
            for p, subj in periods.items():
                if not subj:
                    continue
                if (cls == cls1 and day == day1 and p == p1) or \
                   (cls == cls2 and day == day2 and p == p2):
                    continue
                for t in get_teachers_for_slot(cls, subj):
                    teacher_day_cnt.setdefault(t, {})
                    teacher_day_cnt[t][day] = teacher_day_cnt[t].get(day, 0) + 1
    if subj2:
        for t in get_teachers_for_slot(cls1, subj2):
            teacher_day_cnt.setdefault(t, {})
            teacher_day_cnt[t][day1] = teacher_day_cnt[t].get(day1, 0) + 1
    for t in get_teachers_for_slot(cls2, subj1):
        teacher_day_cnt.setdefault(t, {})
        teacher_day_cnt[t][day2] = teacher_day_cnt[t].get(day2, 0) + 1

    affected = set()
    if subj2:
        for t in get_teachers_for_slot(cls1, subj2):
            affected.add((t, day1))
    for t in get_teachers_for_slot(cls2, subj1):
        affected.add((t, day2))
    for t, day in affected:
        total_today = periods_per_day.get(day, 6)
        unavail_cnt = len(unavailable.get(t, {}).get(day, []))
        available   = total_today - unavail_cnt
        max_ok      = available - 1
        if max_ok <= 0:
            continue
        if teacher_day_cnt.get(t, {}).get(day, 0) > max_ok:
            return False

    # ── R2: 週コマ数保全（異クラス交換時のみ） ────────────────
    # 同クラス内スワップはコマ数が自動的に保たれるため不要。
    # 異クラス・異教科スワップではそれぞれのクラスで教科数が増減するため検証が必要。
    if cls1 != cls2:
        wp = st.session_state["weekly_periods"]

        def _count_subj(cls, subj):
            return sum(
                1 for prs in timetable.get(cls, {}).values()
                for s in prs.values() if s == subj
            )

        # subj1: cls1 から cls2 へ移動
        # cls1 は subj1 を1コマ失う → 必要コマ数を下回らないか
        if _count_subj(cls1, subj1) - 1 < wp.get(cls1, {}).get(subj1, 0):
            return False
        # cls2 は subj1 を1コマ得る → 必要コマ数を超えないか
        if _count_subj(cls2, subj1) + 1 > wp.get(cls2, {}).get(subj1, 0):
            return False

        if subj2:
            # subj2: cls2 から cls1 へ移動
            # cls2 は subj2 を1コマ失う → 必要コマ数を下回らないか
            if _count_subj(cls2, subj2) - 1 < wp.get(cls2, {}).get(subj2, 0):
                return False
            # cls1 は subj2 を1コマ得る → 必要コマ数を超えないか
            if _count_subj(cls1, subj2) + 1 > wp.get(cls1, {}).get(subj2, 0):
                return False

    # ── R8: クラス間教科同期 ──────────────────────────────────
    def after_swap(cls, day, p):
        if cls == cls1 and day == day1 and p == p1:
            return subj2
        if cls == cls2 and day == day2 and p == p2:
            return subj1
        return timetable.get(cls, {}).get(day, {}).get(p)

    for pair in sync_pairs:
        c1_p, s1_p = pair["class1"], pair["subject1"]
        c2_p, s2_p = pair["class2"], pair["subject2"]
        if subj2:
            if cls1 == c1_p and subj2 == s1_p:
                other = after_swap(c2_p, day1, p1)
                if other is not None and other != s2_p:
                    return False
            if cls1 == c2_p and subj2 == s2_p:
                other = after_swap(c1_p, day1, p1)
                if other is not None and other != s1_p:
                    return False
        if cls2 == c1_p and subj1 == s1_p:
            other = after_swap(c2_p, day2, p2)
            if other is not None and other != s2_p:
                return False
        if cls2 == c2_p and subj1 == s2_p:
            other = after_swap(c1_p, day2, p2)
            if other is not None and other != s1_p:
                return False

    return True


def get_swap_violations(
    timetable: dict,
    timetable_sg: dict,
    cls1: str, day1: str, p1: int,
    cls2: str, day2: str, p2: int,
) -> list[str]:
    """
    is_valid_swap と同じ順序で全制約を検査し、
    違反しているものを日本語メッセージのリストで返す。
    空リストなら交換可能。
    """
    violations = []

    if cls1 == cls2 and day1 == day2 and p1 == p2:
        return ["同一スロットが選択されています。"]

    periods_per_day = st.session_state["periods_per_day"]
    unavailable     = st.session_state["unavailable"]
    special_rooms   = st.session_state["special_rooms"]
    sync_pairs      = st.session_state.get("class_subject_sync", [])
    small_group_cls = st.session_state["small_group_classes"]
    wp              = st.session_state["weekly_periods"]

    subj1 = timetable.get(cls1, {}).get(day1, {}).get(p1)
    subj2 = timetable.get(cls2, {}).get(day2, {}).get(p2)

    if not subj1:
        return ["コマAに授業が入っていません。"]
    if cls1 == cls2 and subj1 == subj2:
        return ["同じクラスの同じ教科同士のため入れ替えの意味がありません。"]
    if p1 > periods_per_day.get(day1, 6) or p2 > periods_per_day.get(day2, 6):
        return ["選択した時限がその曜日の授業コマ数を超えています。"]

    same_cls_same_day = (cls1 == cls2 and day1 == day2)

    # ── R4: 教員不在 ──────────────────────────────────────────
    if subj2:
        for t in get_teachers_for_slot(cls1, subj2):
            if p1 in unavailable.get(t, {}).get(day1, []):
                violations.append(
                    f"R4（教員不在）: {t} 先生は {day1}曜{p1}限が不在コマです"
                    f"（{cls1}「{subj2}」を移動できません）"
                )
    for t in get_teachers_for_slot(cls2, subj1):
        if p2 in unavailable.get(t, {}).get(day2, []):
            violations.append(
                f"R4（教員不在）: {t} 先生は {day2}曜{p2}限が不在コマです"
                f"（{cls2}「{subj1}」を移動できません）"
            )

    # ── R5: 同日同教科 ────────────────────────────────────────
    if subj2:
        for p_x, s_x in timetable.get(cls1, {}).get(day1, {}).items():
            if p_x == p1:
                continue
            if same_cls_same_day and p_x == p2:
                continue
            if s_x == subj2:
                violations.append(
                    f"R5（同日同教科）: {cls1} の {day1}曜日に「{subj2}」が既に {p_x}限に入っています"
                )
                break
    for p_x, s_x in timetable.get(cls2, {}).get(day2, {}).items():
        if p_x == p2:
            continue
        if same_cls_same_day and p_x == p1:
            continue
        if s_x == subj1:
            violations.append(
                f"R5（同日同教科）: {cls2} の {day2}曜日に「{subj1}」が既に {p_x}限に入っています"
            )
            break

    # ── R1: 教員重複 ─────────────────────────────────────────
    def busy_at_v(day, p, exclude_cls):
        busy = {}
        for cls, schedule in timetable.items():
            if cls == exclude_cls:
                continue
            s = schedule.get(day, {}).get(p)
            if s:
                for t in get_teachers_for_slot(cls, s):
                    busy[t] = cls
        for sg_name, schedule in timetable_sg.items():
            s = schedule.get(day, {}).get(p)
            if s:
                for t in small_group_cls.get(sg_name, {}).get("teachers", []):
                    busy[t] = sg_name
        return busy

    busy_src = busy_at_v(day1, p1, cls1)
    busy_dst = busy_at_v(day2, p2, cls2)

    t1_set = set(get_teachers_for_slot(cls2, subj1))
    t2_set = set(get_teachers_for_slot(cls1, subj2)) if subj2 else set()

    for t in t2_set:
        if t in busy_src:
            violations.append(
                f"R1（教員重複）: {t} 先生は {day1}曜{p1}限に "
                f"【{busy_src[t]}】で既に授業があります"
                f"（「{subj2}」をここに移動できません）"
            )
    for t in t1_set:
        if t in busy_dst:
            violations.append(
                f"R1（教員重複）: {t} 先生は {day2}曜{p2}限に "
                f"【{busy_dst[t]}】で既に授業があります"
                f"（「{subj1}」をここに移動できません）"
            )

    # ── R3: 特別教室容量 ──────────────────────────────────────
    for room in special_rooms:
        subj_r = room["subject"]
        cap    = room["capacity"]
        if subj2 == subj_r:
            count = sum(
                1 for cls, sched in timetable.items()
                if cls != cls1 and sched.get(day1, {}).get(p1) == subj_r
                and not (cls == cls2 and day1 == day2 and p1 == p2)
            )
            if count + 1 > cap:
                violations.append(
                    f"R3（特別教室）: {day1}曜{p1}限に「{subj_r}」を使う授業が"
                    f"既に {count} クラスあり、上限（{cap} クラス）を超えます"
                )
        if subj1 == subj_r:
            count = sum(
                1 for cls, sched in timetable.items()
                if cls != cls2 and sched.get(day2, {}).get(p2) == subj_r
                and not (cls == cls1 and day1 == day2 and p1 == p2)
            )
            if count + 1 > cap:
                violations.append(
                    f"R3（特別教室）: {day2}曜{p2}限に「{subj_r}」を使う授業が"
                    f"既に {count} クラスあり、上限（{cap} クラス）を超えます"
                )

    # ── R2: 週コマ数 ─────────────────────────────────────────
    if cls1 != cls2:
        def _cnt(cls, subj):
            return sum(1 for prs in timetable.get(cls, {}).values()
                       for s in prs.values() if s == subj)

        if _cnt(cls1, subj1) - 1 < wp.get(cls1, {}).get(subj1, 0):
            violations.append(
                f"R2（週コマ数）: {cls1} の「{subj1}」が"
                f" {_cnt(cls1, subj1) - 1} コマになり"
                f"必要数（{wp.get(cls1, {}).get(subj1, 0)} コマ）を下回ります"
            )
        if _cnt(cls2, subj1) + 1 > wp.get(cls2, {}).get(subj1, 0):
            violations.append(
                f"R2（週コマ数）: {cls2} の「{subj1}」が"
                f" {_cnt(cls2, subj1) + 1} コマになり"
                f"必要数（{wp.get(cls2, {}).get(subj1, 0)} コマ）を超えます"
            )
        if subj2:
            if _cnt(cls2, subj2) - 1 < wp.get(cls2, {}).get(subj2, 0):
                violations.append(
                    f"R2（週コマ数）: {cls2} の「{subj2}」が"
                    f" {_cnt(cls2, subj2) - 1} コマになり"
                    f"必要数（{wp.get(cls2, {}).get(subj2, 0)} コマ）を下回ります"
                )
            if _cnt(cls1, subj2) + 1 > wp.get(cls1, {}).get(subj2, 0):
                violations.append(
                    f"R2（週コマ数）: {cls1} の「{subj2}」が"
                    f" {_cnt(cls1, subj2) + 1} コマになり"
                    f"必要数（{wp.get(cls1, {}).get(subj2, 0)} コマ）を超えます"
                )

    # ── R7: 教員日内コマ数上限 ───────────────────────────────
    teacher_day_cnt: dict = {}
    for cls, schedule in timetable.items():
        for day, periods in schedule.items():
            for p, subj in periods.items():
                if not subj:
                    continue
                if (cls == cls1 and day == day1 and p == p1) or \
                   (cls == cls2 and day == day2 and p == p2):
                    continue
                for t in get_teachers_for_slot(cls, subj):
                    teacher_day_cnt.setdefault(t, {})
                    teacher_day_cnt[t][day] = teacher_day_cnt[t].get(day, 0) + 1
    if subj2:
        for t in get_teachers_for_slot(cls1, subj2):
            teacher_day_cnt.setdefault(t, {})
            teacher_day_cnt[t][day1] = teacher_day_cnt[t].get(day1, 0) + 1
    for t in get_teachers_for_slot(cls2, subj1):
        teacher_day_cnt.setdefault(t, {})
        teacher_day_cnt[t][day2] = teacher_day_cnt[t].get(day2, 0) + 1

    affected = set()
    if subj2:
        for t in get_teachers_for_slot(cls1, subj2):
            affected.add((t, day1))
    for t in get_teachers_for_slot(cls2, subj1):
        affected.add((t, day2))
    for t, day in affected:
        total_today  = periods_per_day.get(day, 6)
        unavail_cnt  = len(unavailable.get(t, {}).get(day, []))
        max_ok       = total_today - unavail_cnt - 1
        if max_ok <= 0:
            continue
        cnt = teacher_day_cnt.get(t, {}).get(day, 0)
        if cnt > max_ok:
            violations.append(
                f"R7（教員日内コマ数）: {t} 先生の {day}曜の授業が"
                f" {cnt} コマになり上限（{max_ok} コマ）を超えます"
            )

    # ── R8: クラス間教科同期 ──────────────────────────────────
    def after_swap_v(cls, day, p):
        if cls == cls1 and day == day1 and p == p1:
            return subj2
        if cls == cls2 and day == day2 and p == p2:
            return subj1
        return timetable.get(cls, {}).get(day, {}).get(p)

    for pair in sync_pairs:
        c1_p, s1_p = pair["class1"], pair["subject1"]
        c2_p, s2_p = pair["class2"], pair["subject2"]
        if subj2:
            if cls1 == c1_p and subj2 == s1_p:
                other = after_swap_v(c2_p, day1, p1)
                if other is not None and other != s2_p:
                    violations.append(
                        f"R8（教科同期）: {c1_p}「{s1_p}」と {c2_p}「{s2_p}」は"
                        f"同時間帯に揃える必要がありますが、{day1}曜{p1}限でずれます"
                    )
            if cls1 == c2_p and subj2 == s2_p:
                other = after_swap_v(c1_p, day1, p1)
                if other is not None and other != s1_p:
                    violations.append(
                        f"R8（教科同期）: {c1_p}「{s1_p}」と {c2_p}「{s2_p}」は"
                        f"同時間帯に揃える必要がありますが、{day1}曜{p1}限でずれます"
                    )
        if cls2 == c1_p and subj1 == s1_p:
            other = after_swap_v(c2_p, day2, p2)
            if other is not None and other != s2_p:
                violations.append(
                    f"R8（教科同期）: {c1_p}「{s1_p}」と {c2_p}「{s2_p}」は"
                    f"同時間帯に揃える必要がありますが、{day2}曜{p2}限でずれます"
                )
        if cls2 == c2_p and subj1 == s2_p:
            other = after_swap_v(c1_p, day2, p2)
            if other is not None and other != s1_p:
                violations.append(
                    f"R8（教科同期）: {c1_p}「{s1_p}」と {c2_p}「{s2_p}」は"
                    f"同時間帯に揃える必要がありますが、{day2}曜{p2}限でずれます"
                )

    return violations


def find_valid_swaps(
    timetable: dict,
    timetable_sg: dict,
    src_cls: str, src_day: str, src_period: int
) -> list:
    """src スロットと交換可能な全スロットを返す。"""
    src_subj = timetable.get(src_cls, {}).get(src_day, {}).get(src_period)
    if not src_subj:
        return []
    periods_per_day = st.session_state["periods_per_day"]
    results = []
    for cls in timetable:
        for day in DAYS:
            day_periods = periods_per_day.get(day, 6)
            for p in range(1, day_periods + 1):
                if cls == src_cls and day == src_day and p == src_period:
                    continue
                if is_valid_swap(timetable, timetable_sg,
                                 src_cls, src_day, src_period,
                                 cls, day, p):
                    tgt_subj = timetable.get(cls, {}).get(day, {}).get(p) or ""
                    results.append((cls, day, p, tgt_subj))
    return results


def get_swap_impact(
    timetable: dict,
    teacher: str,
    cls_a: str, day_a: str, period_a: int,
    cls_b: str, day_b: str, period_b: int,
) -> list:
    """
    (cls_a,day_a,period_a) ↔ (cls_b,day_b,period_b) を入れ替えたとき、
    teacher 以外のどの教員のスケジュールが変化するかを返す。
    戻り値: [{"teacher", "from_cls","from_day","from_period","from_subj",
               "to_cls","to_day","to_period"}, ...]
    """
    subj_a = timetable.get(cls_a, {}).get(day_a, {}).get(period_a)
    subj_b = timetable.get(cls_b, {}).get(day_b, {}).get(period_b)

    impacts = []
    seen = set()

    # subj_a の共同担当者: slot_a → slot_b へ移動
    if subj_a:
        for t in get_teachers_for_slot(cls_a, subj_a):
            if t != teacher and t not in seen:
                seen.add(t)
                impacts.append({
                    "teacher": t,
                    "from_cls": cls_a, "from_day": day_a,
                    "from_period": period_a, "from_subj": subj_a,
                    "to_cls": cls_b, "to_day": day_b, "to_period": period_b,
                })

    # subj_b の共同担当者: slot_b → slot_a へ移動
    if subj_b:
        for t in get_teachers_for_slot(cls_b, subj_b):
            if t != teacher and t not in seen:
                seen.add(t)
                impacts.append({
                    "teacher": t,
                    "from_cls": cls_b, "from_day": day_b,
                    "from_period": period_b, "from_subj": subj_b,
                    "to_cls": cls_a, "to_day": day_a, "to_period": period_a,
                })

    return impacts


# ============================================================
# Part 4: Excel 出力
# ============================================================

def export_timetable_to_excel(
    timetable: dict,
    timetable_sg: dict,
    manual_timetable: dict
) -> bytes:
    """
    生成済み時間割をExcelに出力する。
    ①通常学級シート ②少人数学級シート ③教員別シート
    """
    wb  = openpyxl.Workbook()
    wb.remove(wb.active)
    thin        = Side(style="thin")
    border      = Border(left=thin, right=thin, top=thin, bottom=thin)
    wrap_align  = Alignment(wrap_text=True, horizontal="center", vertical="center")
    max_p       = get_max_periods()

    def make_fill(color):
        return PatternFill("solid", fgColor=color)

    def write_timetable_sheet(ws, cls, schedule, manual, is_sg=False):
        """共通の時間割シート書き込み処理。"""
        # ヘッダー行（曜日）
        ws.append(["時限"] + DAYS)
        _apply_header_style(ws)

        for p in range(1, max_p + 1):
            row_cells = []
            for day in DAYS:
                day_periods = st.session_state["periods_per_day"].get(day, 6)
                if p > day_periods:
                    row_cells.append("─")
                    continue
                subj = schedule.get(day, {}).get(p, "")
                if not subj:
                    row_cells.append("")
                    continue
                if is_sg:
                    teachers = st.session_state["small_group_classes"] \
                        .get(cls, {}).get("teachers", [])
                else:
                    teachers = get_teachers_for_slot(cls, subj)
                cell_val = f"{subj}\n{'・'.join(teachers)}" if teachers else subj
                row_cells.append(cell_val)

            ws.append([f"{p}限"] + row_cells)
            row_idx = p + 1
            ws.row_dimensions[row_idx].height = 35

            for col_idx, day in enumerate(DAYS, start=2):
                cell = ws.cell(row=row_idx, column=col_idx)
                cell.border    = border
                cell.alignment = wrap_align
                day_periods = st.session_state["periods_per_day"].get(day, 6)
                if p > day_periods:
                    continue
                subj = schedule.get(day, {}).get(p, "")
                if not subj:
                    continue
                is_manual = manual.get(cls, {}).get(day, {}).get(p) is not None
                if is_sg:
                    cell.fill = make_fill(COLOR_SG)
                elif is_manual:
                    cell.fill = make_fill(COLOR_MANUAL)
                else:
                    cell.fill = make_fill(COLOR_AUTO)

            # 時限列のスタイル
            lc = ws.cell(row=row_idx, column=1)
            lc.border    = border
            lc.alignment = wrap_align
            lc.fill      = make_fill(COLOR_HEADER)
            lc.font      = Font(bold=True)

        # 列幅設定
        ws.column_dimensions["A"].width = 8
        for col_idx in range(2, len(DAYS) + 2):
            ws.column_dimensions[get_column_letter(col_idx)].width = 18

    # ── ①通常学級シート ──────────────────────────────────────
    for cls in get_all_classes():
        ws = wb.create_sheet(cls)
        write_timetable_sheet(
            ws, cls,
            timetable.get(cls, {}),
            manual_timetable, is_sg=False
        )

    # ── ②少人数学級シート ────────────────────────────────────
    for sg_name in st.session_state["small_group_classes"]:
        ws = wb.create_sheet(sg_name)
        write_timetable_sheet(
            ws, sg_name,
            timetable_sg.get(sg_name, {}),
            {}, is_sg=True
        )

    # ── ③教員別シート ────────────────────────────────────────
    for teacher in st.session_state["teachers"]:
        ws = wb.create_sheet(f"教員_{teacher}")
        ws.append(["時限"] + DAYS)
        _apply_header_style(ws)

        for p in range(1, max_p + 1):
            row_cells = []
            for day in DAYS:
                day_periods = st.session_state["periods_per_day"].get(day, 6)
                if p > day_periods:
                    row_cells.append("─")
                    continue
                if p in st.session_state["unavailable"].get(teacher, {}).get(day, []):
                    row_cells.append("不在")
                    continue
                cell_val = "空き"
                for cls, schedule in timetable.items():
                    subj = schedule.get(day, {}).get(p)
                    if subj and teacher in get_teachers_for_slot(cls, subj):
                        cell_val = f"{cls}\n{subj}"
                        break
                if cell_val == "空き":
                    for sg_name, sg_schedule in timetable_sg.items():
                        subj = sg_schedule.get(day, {}).get(p)
                        if subj and teacher in st.session_state[
                                "small_group_classes"].get(sg_name, {}).get("teachers", []):
                            cell_val = f"{sg_name}\n{subj}"
                            break
                row_cells.append(cell_val)

            ws.append([f"{p}限"] + row_cells)
            row_idx = p + 1
            ws.row_dimensions[row_idx].height = 35

            for col_idx, day in enumerate(DAYS, start=2):
                cell = ws.cell(row=row_idx, column=col_idx)
                cell.border    = border
                cell.alignment = wrap_align
                day_periods = st.session_state["periods_per_day"].get(day, 6)
                if p > day_periods:
                    continue
                val = row_cells[col_idx - 2]
                if val == "不在":
                    cell.fill = make_fill(COLOR_ABSENT)
                elif val == "空き":
                    pass
                else:
                    cell.fill = make_fill(COLOR_AUTO)

            lc = ws.cell(row=row_idx, column=1)
            lc.border = border; lc.alignment = wrap_align
            lc.fill   = make_fill(COLOR_HEADER)
            lc.font   = Font(bold=True)

        ws.column_dimensions["A"].width = 8
        for col_idx in range(2, len(DAYS) + 2):
            ws.column_dimensions[get_column_letter(col_idx)].width = 18

    # ── ④全教員一覧シート ─────────────────────────────────────
    ws = wb.create_sheet("全教員一覧")
    periods_per_day = st.session_state["periods_per_day"]
    unavailable     = st.session_state["unavailable"]
    sg_classes      = st.session_state["small_group_classes"]

    # ヘッダー行（教員名 | 月1 月2 ... 金6）
    col_labels = [
        f"{day}{p}"
        for day in DAYS
        for p in range(1, periods_per_day.get(day, 6) + 1)
    ]
    ws.append(["教員名"] + col_labels)
    _apply_header_style(ws)

    for teacher in st.session_state["teachers"]:
        row_vals = [teacher]
        for day in DAYS:
            day_periods = periods_per_day.get(day, 6)
            for p in range(1, day_periods + 1):
                if p in unavailable.get(teacher, {}).get(day, []):
                    row_vals.append("不在")
                    continue
                cell_val = ""
                for cls, schedule in timetable.items():
                    subj = schedule.get(day, {}).get(p)
                    if subj and teacher in get_teachers_for_slot(cls, subj):
                        cell_val = f"{cls} {subj}"
                        break
                if not cell_val:
                    for sg_name, sg_schedule in timetable_sg.items():
                        subj = sg_schedule.get(day, {}).get(p)
                        if subj and teacher in sg_classes.get(
                                sg_name, {}).get("teachers", []):
                            cell_val = f"{sg_name} {subj}"
                            break
                row_vals.append(cell_val)
        ws.append(row_vals)

    # スタイル適用
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.border    = border
            cell.alignment = wrap_align
            if cell.column == 1:
                cell.fill = make_fill(COLOR_HEADER)
                cell.font = Font(bold=True)
            elif cell.value == "不在":
                cell.fill = make_fill(COLOR_ABSENT)
            elif cell.value:
                cell.fill = make_fill(COLOR_AUTO)
    ws.column_dimensions["A"].width = 14
    for col_idx in range(2, len(col_labels) + 2):
        ws.column_dimensions[get_column_letter(col_idx)].width = 12

    # ── ⑤全クラス一覧シート ───────────────────────────────────
    ws = wb.create_sheet("全クラス一覧")

    # ヘッダー行（クラス名 | 月1 月2 ... 金6）
    col_labels_cls = [
        f"{day}{p}"
        for day in DAYS
        for p in range(1, periods_per_day.get(day, 6) + 1)
    ]
    ws.append(["クラス"] + col_labels_cls)
    _apply_header_style(ws)

    all_cls_list = list(get_all_classes()) + list(st.session_state["small_group_classes"].keys())
    for cls in all_cls_list:
        is_sg = cls in st.session_state["small_group_classes"]
        schedule = timetable_sg.get(cls, {}) if is_sg else timetable.get(cls, {})
        row_vals = [cls]
        for day in DAYS:
            dp = periods_per_day.get(day, 6)
            for p in range(1, dp + 1):
                subj = schedule.get(day, {}).get(p, "")
                if subj:
                    if is_sg:
                        teachers = st.session_state["small_group_classes"] \
                            .get(cls, {}).get("teachers", [])
                    else:
                        teachers = get_teachers_for_slot(cls, subj)
                    cell_val = f"{subj}\n{'・'.join(teachers)}" if teachers else subj
                else:
                    cell_val = ""
                row_vals.append(cell_val)
        ws.append(row_vals)

    # スタイル適用
    for row_idx, row in enumerate(ws.iter_rows(min_row=2), start=2):
        ws.row_dimensions[row_idx].height = 35
        cls_name = ws.cell(row=row_idx, column=1).value or ""
        is_sg_row = cls_name in st.session_state["small_group_classes"]
        col_cursor = 1
        for day in DAYS:
            dp = periods_per_day.get(day, 6)
            for p in range(1, dp + 1):
                col_cursor += 1
                cell = ws.cell(row=row_idx, column=col_cursor)
                cell.border    = border
                cell.alignment = wrap_align
                if cell.value:
                    is_manual = manual_timetable.get(cls_name, {}).get(day, {}).get(p) is not None
                    if is_sg_row:
                        cell.fill = make_fill(COLOR_SG)
                    elif is_manual:
                        cell.fill = make_fill(COLOR_MANUAL)
                    else:
                        cell.fill = make_fill(COLOR_AUTO)
        # クラス名列のスタイル
        lc = ws.cell(row=row_idx, column=1)
        lc.border    = border
        lc.alignment = wrap_align
        lc.fill      = make_fill(COLOR_HEADER)
        lc.font      = Font(bold=True)

    ws.column_dimensions["A"].width = 12
    for col_idx in range(2, len(col_labels_cls) + 2):
        ws.column_dimensions[get_column_letter(col_idx)].width = 14

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ============================================================
# Part 4: 生成画面 render_generate()
# ============================================================

def render_generate():
    st.header("🎲 時間割生成・確認・出力")

    classes = get_all_classes()
    if not classes:
        st.warning("先にSTEP 1〜9を設定してください")
        return

    # ── バリデーション ────────────────────────────────────────
    warnings_list = validate_all_settings()
    if warnings_list:
        with st.expander(
            f"⚠️ 設定に {len(warnings_list)} 件の警告があります（クリックで展開）",
            expanded=False
        ):
            for w in warnings_list:
                st.warning(w)

    # ── ソフト制約設定 ────────────────────────────────────────
    render_soft_constraints()
    st.markdown("---")

    # ── 生成ボタン ────────────────────────────────────────────
    col1, col2, col3 = st.columns([3, 2, 1])
    with col1:
        generate_btn = st.button(
            "🚀 時間割を自動生成する",
            type="primary",
            use_container_width=True
        )
    with col2:
        has_prev = bool(st.session_state.get("generated_timetable"))
        alt_btn = st.button(
            "🔀 別パターンを生成する",
            use_container_width=True,
            disabled=not has_prev,
            help="現在と異なる配置パターンを探します（生成後に有効になります）"
        )
    with col3:
        time_limit = st.number_input(
            "制限（秒）",
            min_value=10, max_value=120, value=30, step=5,
            help="この秒数を超えるとOR-Toolsに切り替えます"
        )

    first_subjects = st.session_state.get("soft_first_subjects", [])
    subject_btn = st.button(
        "📚 教科別段階生成",
        use_container_width=True,
        disabled=not first_subjects,
        help="先に配置する教科を選択してから実行してください（ソフト制約設定で選択）"
    )

    # パターン番号表示
    pattern_no = st.session_state.get("timetable_pattern_no", 0)
    if has_prev:
        label = "標準パターン" if pattern_no == 1 else f"パターン #{pattern_no}"
        st.caption(f"🎲 現在表示中: {label}　　別パターンを生成するたびに番号が増えます")

    # ── 生成処理の共通ロジック ───────────────────────────────
    def _run_generation(forbidden_slots: frozenset = frozenset(), ort_seed: int = 0):
        """
        forbidden_slots: 別パターン生成時に禁止する (cls,day,period,subject) の集合。
                         通常生成時は空集合。
        ort_seed:        OR-Toolsフォールバック時のランダムシード。
        """
        with st.spinner("時間割を生成中..."):
            timetable = {
                cls: {day: {} for day in DAYS}
                for cls in classes
            }
            timetable_sg = {
                sg: {day: {} for day in DAYS}
                for sg in st.session_state["small_group_classes"]
            }
            weekly_periods_sg = {
                sg: sg_data.get("weekly_periods", {})
                for sg, sg_data in st.session_state["small_group_classes"].items()
            }
            weekly_remaining, weekly_remaining_sg = preprocess(
                timetable, timetable_sg,
                st.session_state["manual_timetable"],
                st.session_state["manual_timetable_sg"],
                st.session_state["weekly_periods"],
                weekly_periods_sg
            )
            priority_subjects = st.session_state.get("soft_priority_subjects", [])
            assignments        = st.session_state["assignments"]
            special_rooms      = st.session_state["special_rooms"]
            unavailable        = st.session_state["unavailable"]
            periods_per_day    = st.session_state["periods_per_day"]
            sg_classes         = st.session_state["small_group_classes"]

            result = solve_ortools(
                timetable, timetable_sg,
                weekly_remaining, assignments,
                special_rooms, unavailable,
                periods_per_day, sg_classes,
                priority_subjects,
                random_seed=ort_seed,
                time_limit=float(time_limit)
            )

            used_ortools = True
            used_grade_split = False

            # ── 学年別段階生成フォールバック ──────────────────
            # OR-Toolsが失敗し、複数学年がある場合のみ試みる
            n_grades = len(set(get_grade(c) for c in classes))
            if result != "solved" and n_grades >= 2:
                st.info(
                    "🔀 学年別段階生成を試みます…"
                    "（各学年を順番に生成し、教員の衝突を調整します）"
                )
                # 新しいtimetableを用意してpreprocess
                timetable_gs = {cls: {day: {} for day in DAYS} for cls in classes}
                timetable_sg_gs = {
                    sg: {day: {} for day in DAYS} for sg in timetable_sg
                }
                wr_gs, _ = preprocess(
                    timetable_gs, timetable_sg_gs,
                    st.session_state["manual_timetable"],
                    st.session_state["manual_timetable_sg"],
                    st.session_state["weekly_periods"],
                    weekly_periods_sg
                )
                result = solve_grade_by_grade(
                    timetable_gs, timetable_sg_gs,
                    wr_gs, assignments,
                    special_rooms, unavailable,
                    periods_per_day, sg_classes,
                    priority_subjects,
                    ort_time_limit=float(time_limit) * 2
                )
                if result == "solved":
                    timetable    = timetable_gs
                    timetable_sg = timetable_sg_gs
                    weekly_remaining = wr_gs
                used_grade_split = True

            if result == "solved":
                st.session_state["generated_timetable"]    = timetable
                st.session_state["generated_timetable_sg"] = timetable_sg
                if used_grade_split:
                    method = "学年別段階生成"
                else:
                    method = "OR-Tools CP-SAT"
                pno = st.session_state.get("timetable_pattern_no", 0)
                lbl = "標準パターン" if pno == 1 else f"パターン #{pno}"
                from datetime import datetime as _dt
                history_entry = {
                    "timetable":    timetable,
                    "timetable_sg": timetable_sg,
                    "label":        f"{lbl}（{method}）",
                    "timestamp":    _dt.now().strftime("%H:%M:%S"),
                }
                st.session_state.setdefault("timetable_history", []).append(history_entry)
                st.success(f"✅ {lbl} の生成が完了しました！（{method}）")
                report_r6_violations(timetable)
            else:
                report_conflict(timetable, weekly_remaining)
        return result

    # ── 教科別段階生成 ──────────────────────────────────────────
    def _run_subject_generation():
        """選択した教科を先に配置し、残りを後から配置する段階的生成。"""
        fs = st.session_state.get("soft_first_subjects", [])
        with st.spinner(f"教科別段階生成中… Phase 1: {', '.join(fs)}"):
            timetable = {cls: {day: {} for day in DAYS} for cls in classes}
            timetable_sg = {
                sg: {day: {} for day in DAYS}
                for sg in st.session_state["small_group_classes"]
            }
            weekly_periods_sg = {
                sg: sg_data.get("weekly_periods", {})
                for sg, sg_data in st.session_state["small_group_classes"].items()
            }
            weekly_remaining, _ = preprocess(
                timetable, timetable_sg,
                st.session_state["manual_timetable"],
                st.session_state["manual_timetable_sg"],
                st.session_state["weekly_periods"],
                weekly_periods_sg
            )
            priority_subjects = st.session_state.get("soft_priority_subjects", [])
            assignments     = st.session_state["assignments"]
            special_rooms   = st.session_state["special_rooms"]
            unavailable     = st.session_state["unavailable"]
            periods_per_day = st.session_state["periods_per_day"]
            sg_classes      = st.session_state["small_group_classes"]

            result = solve_subject_by_subject(
                timetable, timetable_sg,
                weekly_remaining, assignments,
                special_rooms, unavailable,
                periods_per_day, sg_classes,
                priority_subjects,
                first_subjects=fs,
                ort_time_limit=float(time_limit) * 2
            )

            if result == "solved":
                st.session_state["generated_timetable"]    = timetable
                st.session_state["generated_timetable_sg"] = timetable_sg
                pno = st.session_state.get("timetable_pattern_no", 0)
                lbl = "標準パターン" if pno == 1 else f"パターン #{pno}"
                from datetime import datetime as _dt
                history_entry = {
                    "timetable":    timetable,
                    "timetable_sg": timetable_sg,
                    "label":        f"{lbl}（教科別段階生成: {', '.join(fs)} → 残り）",
                    "timestamp":    _dt.now().strftime("%H:%M:%S"),
                }
                st.session_state.setdefault("timetable_history", []).append(history_entry)
                st.success(
                    f"✅ {lbl} の生成が完了しました！（教科別段階生成）\n"
                    f"Phase 1: {', '.join(fs)} → Phase 2: 残りの教科"
                )
                report_r6_violations(timetable)
            elif result == "failed_phase1":
                st.error(
                    f"❌ Phase 1（{', '.join(fs)}）の配置に失敗しました。\n"
                    "選択した教科の担当教員・週コマ数・教室設定を確認してください。"
                )
            else:
                st.error(
                    "❌ Phase 2（残りの教科）の配置に失敗しました。\n"
                    "Phase 1 の配置後に残るスロットが不足している可能性があります。"
                )
        return result

    # ── 通常生成 ─────────────────────────────────────────────
    if generate_btn:
        st.session_state["timetable_pattern_no"] = 1
        result = _run_generation()
        if result != "solved":
            return

    # ── 別パターン生成 ────────────────────────────────────────
    if alt_btn:
        import random as _rnd
        prev_tt = st.session_state.get("generated_timetable", {})
        # 前の解から全配置を収集し、約25%をランダムに禁止する
        all_placements = [
            (cls, day, p, subj)
            for cls, day_map in prev_tt.items()
            for day, period_map in day_map.items()
            for p, subj in period_map.items()
            if subj
        ]
        if all_placements:
            n_forbid = max(1, len(all_placements) // 4)
            forbidden = frozenset(_rnd.sample(all_placements, n_forbid))
        else:
            forbidden = frozenset()
        pno = st.session_state.get("timetable_pattern_no", 1) + 1
        st.session_state["timetable_pattern_no"] = pno
        result = _run_generation(
            forbidden_slots=forbidden,
            ort_seed=pno
        )
        if result != "solved":
            # 別パターンが見つからなかった場合はカウンタを戻す
            st.session_state["timetable_pattern_no"] = pno - 1
            return

    # ── 教科別段階生成ボタン ──────────────────────────────────
    if subject_btn:
        st.session_state["timetable_pattern_no"] = 1
        result = _run_subject_generation()
        if result != "solved":
            return

    # ── 生成済み時間割の表示 ──────────────────────────────────
    history = st.session_state.get("timetable_history", [])

    if not history:
        st.info("上の「時間割を自動生成する」ボタンを押してください")
        return

    st.markdown("---")

    # 履歴が複数ある場合は選択UIを表示
    if len(history) > 1:
        history_labels = [
            f"{h['timestamp']}  {h['label']}"
            for h in reversed(history)
        ]
        selected_label = st.selectbox(
            "📋 閲覧する時間割を選択",
            options=history_labels,
            index=0,
            help="過去に生成した時間割と現在の時間割を切り替えて閲覧できます",
        )
        selected_pos = history_labels.index(selected_label)
        selected_entry = list(reversed(history))[selected_pos]
        timetable    = selected_entry["timetable"]
        timetable_sg = selected_entry["timetable_sg"]
        if selected_pos == 0:
            st.caption("✅ 最新の時間割を表示しています")
        else:
            st.caption(f"🕐 過去の時間割を表示中: {selected_entry['label']}")
    else:
        selected_entry = history[-1]
        timetable    = selected_entry["timetable"]
        timetable_sg = selected_entry["timetable_sg"]

    _TAB_LABELS = [
        "📘 クラス別", "📗 少人数学級", "👤 教員別",
        "👥 全教員一覧", "📥 ダウンロード", "✏️ 手動編集"
    ]
    active_tab = st.radio(
        "", _TAB_LABELS, horizontal=True,
        key="tt_active_tab", label_visibility="collapsed"
    )

    st.markdown("---")

    if active_tab == "📘 クラス別":
        selected_cls = st.selectbox(
            "クラスを選択",
            options=get_all_classes(),
            key="disp_cls_select"
        )
        if selected_cls:
            st.subheader(f"📘 {selected_cls} の時間割")
            display_class_timetable(
                selected_cls, timetable,
                st.session_state["manual_timetable"]
            )

    elif active_tab == "📗 少人数学級":
        sg_list = list(st.session_state["small_group_classes"].keys())
        if not sg_list:
            st.info("少人数学級が登録されていません")
        else:
            selected_sg = st.selectbox(
                "少人数学級を選択",
                options=sg_list,
                key="disp_sg_select"
            )
            if selected_sg:
                st.subheader(f"📗 {selected_sg} の時間割")
                display_class_timetable(
                    selected_sg, timetable_sg,
                    st.session_state["manual_timetable_sg"],
                    is_small_group=True
                )

    elif active_tab == "👤 教員別":
        teachers = st.session_state["teachers"]
        if not teachers:
            st.info("教員が登録されていません")
        else:
            selected_t = st.selectbox(
                "教員を選択",
                options=teachers,
                key="disp_teacher_select"
            )
            if selected_t:
                st.subheader(f"👤 {selected_t} の時間割")
                display_teacher_timetable(selected_t, timetable, timetable_sg)

                # ── コマの変更確認 ────────────────────────────────
                st.markdown("---")
                st.markdown("#### 🔍 コマの変更確認")
                chk_col1, chk_col2 = st.columns(2)
                with chk_col1:
                    chk_day = st.selectbox("曜日", DAYS, key="chk_day_sel")
                with chk_col2:
                    chk_max_p = st.session_state["periods_per_day"].get(chk_day, 6)
                    chk_period = st.selectbox(
                        "時限",
                        list(range(1, chk_max_p + 1)),
                        format_func=lambda x: f"{x}限",
                        key="chk_period_sel"
                    )

                if st.button("🔍 変更可能か確認する", key="btn_chk_swap"):
                    slot_cls = slot_subj = None
                    for cls, schedule in timetable.items():
                        subj = schedule.get(chk_day, {}).get(chk_period)
                        if subj and selected_t in get_teachers_for_slot(cls, subj):
                            slot_cls, slot_subj = cls, subj
                            break

                    if slot_subj:
                        # 授業コマ: 全コマとのスワップ候補を検索
                        with st.spinner("全制約をチェック中..."):
                            cands = find_valid_swaps(
                                timetable, timetable_sg, slot_cls, chk_day, chk_period
                            )
                        st.session_state["chk_result"] = {
                            "teacher": selected_t, "day": chk_day, "period": chk_period,
                            "type": "occupied", "cls": slot_cls, "subj": slot_subj,
                            "candidates": cands,
                        }
                    else:
                        # 空きコマ: この教員の他コマで同クラス内移動が可能なものを検索
                        with st.spinner("全制約をチェック中..."):
                            cands = []
                            for cls, schedule in timetable.items():
                                for day, periods in schedule.items():
                                    for p, subj in periods.items():
                                        if not subj:
                                            continue
                                        if selected_t not in get_teachers_for_slot(cls, subj):
                                            continue
                                        if day == chk_day and p == chk_period:
                                            continue
                                        if is_valid_swap(
                                            timetable, timetable_sg,
                                            cls, day, p, cls, chk_day, chk_period
                                        ):
                                            cands.append((cls, day, p, subj))
                        st.session_state["chk_result"] = {
                            "teacher": selected_t, "day": chk_day, "period": chk_period,
                            "type": "empty", "candidates": cands,
                        }

                # 結果表示（前回の検索結果が現在の選択と一致する場合のみ）
                chk_result = st.session_state.get("chk_result")
                if (chk_result
                        and chk_result.get("teacher") == selected_t
                        and chk_result.get("day") == chk_day
                        and chk_result.get("period") == chk_period):

                    cands = chk_result["candidates"]

                    if chk_result["type"] == "occupied":
                        cls   = chk_result["cls"]
                        subj  = chk_result["subj"]
                        st.markdown(
                            f"**対象コマ:** {cls} {chk_day}曜 {chk_period}限 ─ "
                            f"**:blue[{subj}]**（担当: {selected_t}）"
                        )
                        # 他の教員が担当するコマ・空きコマのみに絞り込む
                        show_cands = [
                            (c, d, p, s) for c, d, p, s in cands
                            if not s or selected_t not in get_teachers_for_slot(c, s)
                        ]
                        if not show_cands:
                            st.error("❌ このコマは交換できません。")
                        else:
                            st.success(f"✅ {len(show_cands)} 件の交換候補があります。")
                            labels = []
                            for c, d, p, s in show_cands:
                                t_list = get_teachers_for_slot(c, s) if s else []
                                t_str  = "・".join(t_list) if t_list else "─"
                                labels.append(
                                    f"【{c}】{d}曜{p}限「{s if s else '空き'}」"
                                    f" 担当:{t_str}"
                                )
                            sel_label = st.selectbox(
                                "交換候補を選択", labels, key="chk_cand_sel"
                            )
                            sel_idx = labels.index(sel_label)
                            tgt_c, tgt_d, tgt_p, tgt_s = show_cands[sel_idx]
                            if st.button(
                                "✏️ この候補で手動編集タブへ", key="btn_goto_edit_from_chk"
                            ):
                                st.session_state["tt_active_tab"]    = "✏️ 手動編集"
                                st.session_state["edit_select_mode"] = "クラスから選ぶ"
                                st.session_state["edit_cls_select"]  = cls
                                st.session_state["edit_cls_last"]    = cls
                                st.session_state["edit_cls_slot_a"]  = (chk_day, chk_period, cls, subj)
                                st.session_state["edit_cls_slot_b"]  = (tgt_d, tgt_p, tgt_c, tgt_s)
                                st.session_state["edit_cls_cands"]   = None
                                st.rerun()

                    else:  # empty
                        st.markdown(
                            f"**対象コマ:** {selected_t} {chk_day}曜 {chk_period}限 ─ "
                            f"**空きコマ**"
                        )
                        if not cands:
                            st.error("❌ このコマに移動できる授業がありません。")
                        else:
                            st.success(f"✅ {len(cands)} 件の授業をこのコマに移動できます。")
                            labels = []
                            for c, d, p, s in cands:
                                labels.append(
                                    f"【{c}】{d}曜{p}限「{s}」→ {chk_day}曜{chk_period}限に移動"
                                )
                            sel_label = st.selectbox(
                                "移動候補を選択", labels, key="chk_cand_sel"
                            )
                            sel_idx = labels.index(sel_label)
                            src_c, src_d, src_p, src_s = cands[sel_idx]
                            if st.button(
                                "✏️ この候補で手動編集タブへ", key="btn_goto_edit_from_chk"
                            ):
                                st.session_state["tt_active_tab"]    = "✏️ 手動編集"
                                st.session_state["edit_select_mode"] = "クラスから選ぶ"
                                st.session_state["edit_cls_select"]  = src_c
                                st.session_state["edit_cls_last"]    = src_c
                                st.session_state["edit_cls_slot_a"]  = (src_d, src_p, src_c, src_s)
                                st.session_state["edit_cls_slot_b"]  = (chk_day, chk_period, src_c, None)
                                st.session_state["edit_cls_cands"]   = None
                                st.rerun()

                # ── 2コマの入れ替え確認 ────────────────────────────────
                st.markdown("---")
                st.markdown("#### 🔄 同一教員の2コマ入れ替え確認")
                st.caption(
                    "コマをクリックして A・B を選択 → 「確認する」で入れ替え可否と他教員への影響を表示します。"
                )

                # 教員が変わったら選択をリセット
                if st.session_state.get("swap2_last_teacher") != selected_t:
                    st.session_state["swap2_slot_a"] = None
                    st.session_state["swap2_slot_b"] = None
                    st.session_state["swap2_result"] = None
                    st.session_state["swap2_last_teacher"] = selected_t

                _slot_a = st.session_state.get("swap2_slot_a")  # (day, period, cls, subj)
                _slot_b = st.session_state.get("swap2_slot_b")

                # 選択中コマの表示
                _disp_col1, _disp_col2 = st.columns(2)
                with _disp_col1:
                    if _slot_a:
                        _la = (f"【{_slot_a[2]}】{_slot_a[0]}曜{_slot_a[1]}限「{_slot_a[3]}」"
                               if _slot_a[3] else f"{_slot_a[0]}曜{_slot_a[1]}限（空き）")
                        st.info(f"🅐 コマA: {_la}")
                    else:
                        st.info("🅐 コマA: 未選択")
                with _disp_col2:
                    if _slot_b:
                        _lb = (f"【{_slot_b[2]}】{_slot_b[0]}曜{_slot_b[1]}限「{_slot_b[3]}」"
                               if _slot_b[3] else f"{_slot_b[0]}曜{_slot_b[1]}限（空き）")
                        st.info(f"🅑 コマB: {_lb}")
                    else:
                        st.info("🅑 コマB: 未選択")

                # クリック可能な時間割グリッド
                _ppd  = st.session_state["periods_per_day"]
                _maxp = max(_ppd.values()) if _ppd else 6

                _hdr_cols = st.columns([0.6] + [1] * len(DAYS))
                with _hdr_cols[0]:
                    st.markdown("**時限**")
                for _hi, _hd in enumerate(DAYS):
                    with _hdr_cols[_hi + 1]:
                        st.markdown(f"**{_hd}曜**")

                for _p in range(1, _maxp + 1):
                    _row = st.columns([0.6] + [1] * len(DAYS))
                    with _row[0]:
                        st.markdown(f"**{_p}限**")
                    for _di, _d in enumerate(DAYS):
                        with _row[_di + 1]:
                            if _p > _ppd.get(_d, 6):
                                st.write("─")
                                continue
                            # この先生のコマを解決
                            _sc = _ss = None
                            for _c, _sched in timetable.items():
                                _s = _sched.get(_d, {}).get(_p)
                                if _s and selected_t in get_teachers_for_slot(_c, _s):
                                    _sc, _ss = _c, _s
                                    break
                            _is_a = bool(_slot_a and _slot_a[0] == _d and _slot_a[1] == _p)
                            _is_b = bool(_slot_b and _slot_b[0] == _d and _slot_b[1] == _p)
                            _mark  = "🅐 " if _is_a else ("🅑 " if _is_b else "")
                            _label = f"{_mark}{_ss}" if _ss else f"{_mark}空き"
                            if _sc:
                                _label += f"\n({_sc})"
                            if st.button(
                                _label,
                                key=f"s2_cell_{_d}_{_p}",
                                type="primary" if (_is_a or _is_b) else "secondary",
                                use_container_width=True,
                            ):
                                if _is_a:
                                    st.session_state["swap2_slot_a"] = None
                                elif _is_b:
                                    st.session_state["swap2_slot_b"] = None
                                elif not _slot_a:
                                    st.session_state["swap2_slot_a"] = (_d, _p, _sc, _ss)
                                elif not _slot_b:
                                    st.session_state["swap2_slot_b"] = (_d, _p, _sc, _ss)
                                else:
                                    # 両方選択済み → Aを置き換えてBはリセット
                                    st.session_state["swap2_slot_a"] = (_d, _p, _sc, _ss)
                                    st.session_state["swap2_slot_b"] = None
                                st.session_state["swap2_result"] = None
                                st.rerun()

                # 確認ボタン（A・B 両方選択済みのときだけ表示）
                if _slot_a and _slot_b:
                    if st.button("🔄 入れ替え可能か確認する", key="btn_chk_swap2"):
                        st.session_state["swap2_result"] = {
                            "teacher":  selected_t,
                            "day_a":    _slot_a[0], "period_a": _slot_a[1],
                            "cls_a":    _slot_a[2], "subj_a":   _slot_a[3],
                            "day_b":    _slot_b[0], "period_b": _slot_b[1],
                            "cls_b":    _slot_b[2], "subj_b":   _slot_b[3],
                        }
                else:
                    st.caption("↑ コマを2つクリックして選択してください")

                # 結果表示
                swap2_result = st.session_state.get("swap2_result")
                _slot_a = st.session_state.get("swap2_slot_a")  # 再取得（rerun後）
                _slot_b = st.session_state.get("swap2_slot_b")
                if (swap2_result
                        and swap2_result.get("teacher") == selected_t
                        and _slot_a and _slot_b
                        and swap2_result.get("day_a") == _slot_a[0]
                        and swap2_result.get("period_a") == _slot_a[1]
                        and swap2_result.get("day_b") == _slot_b[0]
                        and swap2_result.get("period_b") == _slot_b[1]):

                    r2 = swap2_result
                    s2_day_a    = r2["day_a"];    s2_period_a = r2["period_a"]
                    s2_day_b    = r2["day_b"];    s2_period_b = r2["period_b"]
                    cls_a2  = r2["cls_a"];  subj_a2 = r2["subj_a"]
                    cls_b2  = r2["cls_b"];  subj_b2 = r2["subj_b"]

                    label_a2 = (
                        f"【{cls_a2}】{s2_day_a}曜{s2_period_a}限「{subj_a2}」"
                        if subj_a2 else f"{s2_day_a}曜{s2_period_a}限（空きコマ）"
                    )
                    label_b2 = (
                        f"【{cls_b2}】{s2_day_b}曜{s2_period_b}限「{subj_b2}」"
                        if subj_b2 else f"{s2_day_b}曜{s2_period_b}限（空きコマ）"
                    )
                    st.markdown(f"**コマ A:** {label_a2}")
                    st.markdown(f"**コマ B:** {label_b2}")

                    # 同一スロット判定
                    if s2_day_a == s2_day_b and s2_period_a == s2_period_b:
                        st.error("❌ 同じコマが選択されています。")
                    elif not subj_a2 and not subj_b2:
                        st.warning("⚠️ 両方とも空きコマです。入れ替えの意味がありません。")
                    else:
                        # 入れ替え後の実効クラスを決定
                        if subj_a2 and subj_b2:
                            # 両方授業: クラスをまたいで交換
                            eff_cls_a2, eff_cls_b2 = cls_a2, cls_b2
                            _vp2 = (cls_a2, s2_day_a, s2_period_a,
                                    cls_b2, s2_day_b, s2_period_b)
                            ok2 = is_valid_swap(timetable, timetable_sg, *_vp2)
                        elif subj_a2:
                            # Aのみ授業: Aを同クラス内でBの時間に移動
                            eff_cls_a2, eff_cls_b2 = cls_a2, cls_a2
                            _vp2 = (cls_a2, s2_day_a, s2_period_a,
                                    cls_a2, s2_day_b, s2_period_b)
                            ok2 = is_valid_swap(timetable, timetable_sg, *_vp2)
                        else:
                            # Bのみ授業: Bを同クラス内でAの時間に移動
                            eff_cls_a2, eff_cls_b2 = cls_b2, cls_b2
                            _vp2 = (cls_b2, s2_day_b, s2_period_b,
                                    cls_b2, s2_day_a, s2_period_a)
                            ok2 = is_valid_swap(timetable, timetable_sg, *_vp2)

                        if not ok2:
                            _viols2 = get_swap_violations(
                                timetable, timetable_sg, *_vp2
                            )
                            st.error("❌ このコマは入れ替えできません。")
                            for _v in _viols2:
                                st.write(f"　・{_v}")
                        else:
                            st.success("✅ このコマは入れ替えできます。")

                            # 他の教員への影響
                            impacts2 = get_swap_impact(
                                timetable, selected_t,
                                eff_cls_a2, s2_day_a, s2_period_a,
                                eff_cls_b2, s2_day_b, s2_period_b,
                            )
                            if not impacts2:
                                st.info("他の教員への影響はありません。")
                            else:
                                st.markdown("**他の教員への影響:**")
                                for imp in impacts2:
                                    to_subj_label = imp.get("from_subj", "─")
                                    st.write(
                                        f"・**{imp['teacher']}**: "
                                        f"【{imp['from_cls']}】"
                                        f"{imp['from_day']}曜{imp['from_period']}限"
                                        f"「{to_subj_label}」"
                                        f" → "
                                        f"【{imp['to_cls']}】"
                                        f"{imp['to_day']}曜{imp['to_period']}限"
                                    )

                            if st.button("✅ 入れ替えを実行する", key="btn_exec_swap2"):
                                # undo スナップショット
                                snapshot2 = {
                                    c: {d: dict(prs) for d, prs in s.items()}
                                    for c, s in timetable.items()
                                }
                                st.session_state["swap_undo_stack"].append({
                                    "snapshot": snapshot2,
                                    "description": (
                                        f"{selected_t}: {label_a2} ↔ {label_b2}"
                                    ),
                                })
                                # 実際の入れ替え（必ずswapにして既存授業を消さない）
                                if subj_a2 and subj_b2:
                                    # 両方授業: クラスをまたいで交換
                                    timetable[cls_a2][s2_day_a][s2_period_a] = subj_b2
                                    timetable[cls_b2][s2_day_b][s2_period_b] = subj_a2
                                elif subj_a2:
                                    # Aのみ授業: 同クラス内でBの時間と入れ替え
                                    # (Bに別の授業があれば、それがAに来る)
                                    _exist_b = timetable.get(cls_a2, {}).get(s2_day_b, {}).get(s2_period_b)
                                    timetable[cls_a2][s2_day_a][s2_period_a] = _exist_b
                                    timetable[cls_a2].setdefault(s2_day_b, {})[s2_period_b] = subj_a2
                                else:
                                    # Bのみ授業: 同クラス内でAの時間と入れ替え
                                    # (Aに別の授業があれば、それがBに来る)
                                    _exist_a = timetable.get(cls_b2, {}).get(s2_day_a, {}).get(s2_period_a)
                                    timetable[cls_b2][s2_day_b][s2_period_b] = _exist_a
                                    timetable[cls_b2].setdefault(s2_day_a, {})[s2_period_a] = subj_b2
                                st.session_state["swap2_result"] = None
                                st.rerun()

    elif active_tab == "👥 全教員一覧":
        teachers = st.session_state["teachers"]
        if not teachers:
            st.info("教員が登録されていません")
        else:
            st.subheader("👥 全教員の時間割一覧")
            display_all_teachers_timetable(timetable, timetable_sg)

    elif active_tab == "📥 ダウンロード":
        st.markdown("#### 📥 時間割をExcelでダウンロード")
        fname = f"時間割_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        st.download_button(
            label="📥 時間割Excelをダウンロード",
            data=export_timetable_to_excel(
                timetable, timetable_sg,
                st.session_state["manual_timetable"]
            ),
            file_name=fname,
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            use_container_width=True
        )
        st.caption("クラス別・少人数学級・教員別シートが含まれます")

    elif active_tab == "✏️ 手動編集":
        st.subheader("✏️ コマの手動交換")
        st.info(
            "交換したいコマを選択し「候補を検索」を押すと、"
            "全ての制約を満たす交換可能なコマが一覧表示されます。"
        )

        # ── 選択方法 ──────────────────────────────────────────
        select_mode = st.radio(
            "コマの選択方法",
            ["クラスから選ぶ", "教員から選ぶ"],
            horizontal=True,
            key="edit_select_mode"
        )

        edit_cls    = None
        edit_day    = None
        edit_period = None

        if select_mode == "クラスから選ぶ":
            edit_cls = st.selectbox(
                "クラス", get_all_classes(), key="edit_cls_select"
            )

            # クラスが変わったら選択をリセット
            if st.session_state.get("edit_cls_last") != edit_cls:
                st.session_state["edit_cls_slot_a"]  = None
                st.session_state["edit_cls_slot_b"]  = None
                st.session_state["edit_cls_cands"]   = None
                st.session_state["edit_cls_last"]    = edit_cls

            _ec_sa = st.session_state.get("edit_cls_slot_a")  # (day, period, cls, subj) or None
            _ec_sb = st.session_state.get("edit_cls_slot_b")  # (day, period, cls, subj) or None

            # スロットAが選択済みでBが未選択なら候補を自動検索（キャッシュ）
            _ec_cand_set = set()  # (day, period) の集合（同クラス内）
            if _ec_sa and not _ec_sb:
                _ec_sa_day, _ec_sa_per, _ec_sa_cls, _ec_sa_subj = _ec_sa
                if st.session_state.get("edit_cls_cands") is None:
                    with st.spinner("候補を検索中..."):
                        if _ec_sa_subj:
                            _ec_all_cands = find_valid_swaps(
                                timetable, timetable_sg, _ec_sa_cls, _ec_sa_day, _ec_sa_per
                            )
                        else:
                            # 空きコマ選択: このクラスのどの授業をここへ移動できるか
                            _ec_all_cands = []
                            for _d, _prs in timetable.get(edit_cls, {}).items():
                                for _p, _s in _prs.items():
                                    if not _s:
                                        continue
                                    if _d == _ec_sa_day and _p == _ec_sa_per:
                                        continue
                                    if is_valid_swap(
                                        timetable, timetable_sg,
                                        edit_cls, _d, _p,
                                        edit_cls, _ec_sa_day, _ec_sa_per
                                    ):
                                        _ec_all_cands.append((edit_cls, _d, _p, _s))
                    st.session_state["edit_cls_cands"] = _ec_all_cands
                _ec_all_cands = st.session_state.get("edit_cls_cands", [])
                _ec_cand_set  = {(d, p) for c, d, p, s in _ec_all_cands if c == edit_cls}

            # A・B 選択状況の表示
            _ec_col1, _ec_col2 = st.columns(2)
            with _ec_col1:
                if _ec_sa:
                    _ec_sa_lbl = (f"{_ec_sa[0]}曜{_ec_sa[1]}限「{_ec_sa[3]}」"
                                  if _ec_sa[3] else f"{_ec_sa[0]}曜{_ec_sa[1]}限（空き）")
                    _ec_hint = "（ここへ移動できる授業）" if not _ec_sa[3] else "（交換可能）"
                    if _ec_cand_set:
                        st.info(f"🅐 コマA: {_ec_sa_lbl}　🟢 {len(_ec_cand_set)} 件{_ec_hint}")
                    else:
                        st.info(f"🅐 コマA: {_ec_sa_lbl}")
                else:
                    st.info("🅐 コマA: 未選択")
            with _ec_col2:
                if _ec_sb:
                    _ec_sb_lbl = (f"{_ec_sb[0]}曜{_ec_sb[1]}限「{_ec_sb[3]}」"
                                  if _ec_sb[3] else f"{_ec_sb[0]}曜{_ec_sb[1]}限（空き）")
                    st.info(f"🅑 コマB: {_ec_sb_lbl}")
                else:
                    _b_hint2 = "🟢のコマを選択" if _ec_cand_set else "未選択"
                    st.info(f"🅑 コマB: {_b_hint2}")

            # 候補ボタン用CSS（:has() は Chrome 105+ / Firefox 121+ で動作）
            st.markdown("""
<style>
div[data-testid="stMarkdown"]:has(.ec-swap-candidate)
  + div[data-testid="stButton"] button {
    background-color: #c8f7c5 !important;
    border: 2px solid #28a745 !important;
    color: #155724 !important;
}
</style>
""", unsafe_allow_html=True)

            # クリック可能な時間割グリッド
            _ppd_c  = st.session_state["periods_per_day"]
            _maxp_c = max(_ppd_c.values()) if _ppd_c else 6

            _hdr_c = st.columns([0.6] + [1] * len(DAYS))
            with _hdr_c[0]:
                st.markdown("**時限**")
            for _hi_c, _hd_c in enumerate(DAYS):
                with _hdr_c[_hi_c + 1]:
                    st.markdown(f"**{_hd_c}曜**")

            for _p_c in range(1, _maxp_c + 1):
                _row_c = st.columns([0.6] + [1] * len(DAYS))
                with _row_c[0]:
                    st.markdown(f"**{_p_c}限**")
                for _di_c, _d_c in enumerate(DAYS):
                    with _row_c[_di_c + 1]:
                        if _p_c > _ppd_c.get(_d_c, 6):
                            st.write("─")
                            continue
                        _ss_c    = timetable.get(edit_cls, {}).get(_d_c, {}).get(_p_c)
                        _is_a_c  = bool(_ec_sa and _ec_sa[0] == _d_c and _ec_sa[1] == _p_c)
                        _is_b_c  = bool(_ec_sb and _ec_sb[0] == _d_c and _ec_sb[1] == _p_c)
                        _is_cc   = ((_d_c, _p_c) in _ec_cand_set and not _is_a_c and not _is_b_c)
                        if _is_cc:
                            st.markdown(
                                '<div class="ec-swap-candidate"></div>',
                                unsafe_allow_html=True
                            )
                        _mark_c = "🅐 " if _is_a_c else ("🅑 " if _is_b_c else ("🔄 " if _is_cc else ""))
                        _lbl_c  = f"{_mark_c}{_ss_c}" if _ss_c else f"{_mark_c}空き"
                        if st.button(
                            _lbl_c,
                            key=f"ec_cell_{_d_c}_{_p_c}",
                            type="primary" if (_is_a_c or _is_b_c) else "secondary",
                            use_container_width=True,
                        ):
                            if _is_a_c:
                                st.session_state["edit_cls_slot_a"] = None
                                st.session_state["edit_cls_cands"]  = None
                            elif _is_b_c:
                                st.session_state["edit_cls_slot_b"] = None
                            elif not _ec_sa:
                                st.session_state["edit_cls_slot_a"] = (_d_c, _p_c, edit_cls, _ss_c)
                                st.session_state["edit_cls_cands"]  = None
                            elif not _ec_sb:
                                st.session_state["edit_cls_slot_b"] = (_d_c, _p_c, edit_cls, _ss_c)
                            else:
                                # 両方選択済みで別のコマをクリック → Aを置き換え
                                st.session_state["edit_cls_slot_a"] = (_d_c, _p_c, edit_cls, _ss_c)
                                st.session_state["edit_cls_slot_b"] = None
                                st.session_state["edit_cls_cands"]  = None
                            st.rerun()

            # A・B 両方選択済みなら即座に結果を表示
            _ec_sa = st.session_state.get("edit_cls_slot_a")
            _ec_sb = st.session_state.get("edit_cls_slot_b")
            if not (_ec_sa and _ec_sb):
                st.caption("↑ コマを2つクリックして選択してください")
            else:
                _ec_da, _ec_pa, _ec_ca, _ec_ss_a = _ec_sa
                _ec_db, _ec_pb, _ec_cb, _ec_ss_b = _ec_sb

                _ec_la = (f"【{_ec_ca}】{_ec_da}曜{_ec_pa}限「{_ec_ss_a}」"
                          if _ec_ss_a else f"{_ec_da}曜{_ec_pa}限（空きコマ）")
                _ec_lb = (f"【{_ec_cb}】{_ec_db}曜{_ec_pb}限「{_ec_ss_b}」"
                          if _ec_ss_b else f"{_ec_db}曜{_ec_pb}限（空きコマ）")

                st.markdown("---")
                if _ec_da == _ec_db and _ec_pa == _ec_pb and _ec_ca == _ec_cb:
                    st.error("❌ 同じコマが選択されています。")
                elif not _ec_ss_a and not _ec_ss_b:
                    st.warning("⚠️ 両方とも空きコマです。入れ替えの意味がありません。")
                else:
                    # 実効クラスと検証パラメータを決定
                    if _ec_ss_a and _ec_ss_b:
                        _ec_vp = (_ec_ca, _ec_da, _ec_pa, _ec_cb, _ec_db, _ec_pb)
                        _ec_eff_ca, _ec_eff_cb = _ec_ca, _ec_cb
                    elif _ec_ss_a:
                        _ec_vp = (_ec_ca, _ec_da, _ec_pa, _ec_ca, _ec_db, _ec_pb)
                        _ec_eff_ca = _ec_eff_cb = _ec_ca
                    else:
                        _ec_vp = (_ec_cb, _ec_db, _ec_pb, _ec_cb, _ec_da, _ec_pa)
                        _ec_eff_ca = _ec_eff_cb = _ec_cb

                    _ec_ok = is_valid_swap(timetable, timetable_sg, *_ec_vp)

                    if not _ec_ok:
                        _ec_viols = get_swap_violations(timetable, timetable_sg, *_ec_vp)
                        st.error("❌ このコマは入れ替えできません。")
                        for _v in _ec_viols:
                            st.write(f"　・{_v}")
                    else:
                        # 他の教員への影響（クラスから選ぶは教員を特定しないので None 扱い）
                        _ec_impacts = get_swap_impact(
                            timetable, None,
                            _ec_eff_ca, _ec_da, _ec_pa,
                            _ec_eff_cb, _ec_db, _ec_pb,
                        )
                        if not _ec_impacts:
                            st.info("他の教員への影響はありません。")
                        else:
                            st.markdown("**他の教員への影響:**")
                            for _imp in _ec_impacts:
                                st.write(
                                    f"・**{_imp['teacher']}**: "
                                    f"【{_imp['from_cls']}】"
                                    f"{_imp['from_day']}曜{_imp['from_period']}限"
                                    f"「{_imp['from_subj']}」"
                                    f" → "
                                    f"【{_imp['to_cls']}】"
                                    f"{_imp['to_day']}曜{_imp['to_period']}限"
                                )

                        if st.button("✅ 入れ替えを実行する", type="primary",
                                     key="btn_exec_edit_cls"):
                            _ec_snap = {
                                c: {d: dict(prs) for d, prs in s.items()}
                                for c, s in timetable.items()
                            }
                            st.session_state["swap_undo_stack"].append({
                                "snapshot": _ec_snap,
                                "description": f"{_ec_la} ↔ {_ec_lb}",
                            })
                            if _ec_ss_a and _ec_ss_b:
                                timetable[_ec_ca][_ec_da][_ec_pa] = _ec_ss_b
                                timetable[_ec_cb][_ec_db][_ec_pb] = _ec_ss_a
                            elif _ec_ss_a:
                                _ec_ex = timetable.get(_ec_ca, {}).get(_ec_db, {}).get(_ec_pb)
                                timetable[_ec_ca][_ec_da][_ec_pa] = _ec_ex
                                timetable[_ec_ca].setdefault(_ec_db, {})[_ec_pb] = _ec_ss_a
                            else:
                                _ec_ex = timetable.get(_ec_cb, {}).get(_ec_da, {}).get(_ec_pa)
                                timetable[_ec_cb][_ec_db][_ec_pb] = _ec_ex
                                timetable[_ec_cb].setdefault(_ec_da, {})[_ec_pa] = _ec_ss_b
                            st.session_state["generated_timetable"] = timetable
                            st.session_state["edit_cls_slot_a"] = None
                            st.session_state["edit_cls_slot_b"] = None
                            st.session_state["edit_cls_cands"]  = None
                            st.rerun()

        else:  # 教員から選ぶ ── 同一教員の2コマ入れ替え
            teachers_list = st.session_state["teachers"]
            if not teachers_list:
                st.info("教員が登録されていません")
            else:
                edit_teacher = st.selectbox(
                    "教員", teachers_list, key="edit_teacher_select"
                )
                st.caption(
                    "コマをクリックして A・B を選択し、「確認する」で入れ替え可否と"
                    "他教員への影響を確認できます。"
                )

                # 教員が変わったら選択をリセット
                if st.session_state.get("edit2_last_teacher") != edit_teacher:
                    st.session_state["edit2_slot_a"] = None
                    st.session_state["edit2_slot_b"] = None
                    st.session_state["edit2_result"] = None
                    st.session_state["edit2_cands"]  = None
                    st.session_state["edit2_last_teacher"] = edit_teacher

                _e2_slot_a = st.session_state.get("edit2_slot_a")
                _e2_slot_b = st.session_state.get("edit2_slot_b")

                # スロットAが選択済みなら候補を自動検索（キャッシュ）
                _e2_cand_set = set()  # 候補の (day, period) 集合
                if _e2_slot_a and not _e2_slot_b:
                    if st.session_state.get("edit2_cands") is None:
                        with st.spinner("候補を検索中..."):
                            if _e2_slot_a[3]:
                                # 授業コマ選択: 全交換候補を検索
                                _a_cls = _e2_slot_a[2]
                                if _a_cls:
                                    _e2_all_cands = find_valid_swaps(
                                        timetable, timetable_sg,
                                        _a_cls, _e2_slot_a[0], _e2_slot_a[1]
                                    )
                                else:
                                    _e2_all_cands = []
                            else:
                                # 空きコマ選択: この教員のどの授業をここへ移動できるか
                                _e2_all_cands = []
                                for _cls, _sched in timetable.items():
                                    for _d, _prs in _sched.items():
                                        for _p, _s in _prs.items():
                                            if not _s:
                                                continue
                                            if edit_teacher not in get_teachers_for_slot(_cls, _s):
                                                continue
                                            if _d == _e2_slot_a[0] and _p == _e2_slot_a[1]:
                                                continue
                                            if is_valid_swap(
                                                timetable, timetable_sg,
                                                _cls, _d, _p,
                                                _cls, _e2_slot_a[0], _e2_slot_a[1]
                                            ):
                                                _e2_all_cands.append((_cls, _d, _p, _s))
                        st.session_state["edit2_cands"] = _e2_all_cands
                    _e2_cand_set = {(d, p) for _, d, p, _ in
                                    st.session_state.get("edit2_cands", [])}

                # 選択中コマの表示
                _e2_c1, _e2_c2 = st.columns(2)
                with _e2_c1:
                    if _e2_slot_a:
                        _la = (f"【{_e2_slot_a[2]}】{_e2_slot_a[0]}曜{_e2_slot_a[1]}限"
                               f"「{_e2_slot_a[3]}」" if _e2_slot_a[3]
                               else f"{_e2_slot_a[0]}曜{_e2_slot_a[1]}限（空き）")
                        _hint = ("（ここへ移動できる授業）" if not _e2_slot_a[3]
                                 else "（交換可能）")
                        if _e2_cand_set:
                            st.info(f"🅐 コマA: {_la}　🟢 {len(_e2_cand_set)} 件{_hint}")
                        else:
                            st.info(f"🅐 コマA: {_la}")
                    else:
                        st.info("🅐 コマA: 未選択")
                with _e2_c2:
                    if _e2_slot_b:
                        _lb = (f"【{_e2_slot_b[2]}】{_e2_slot_b[0]}曜{_e2_slot_b[1]}限"
                               f"「{_e2_slot_b[3]}」" if _e2_slot_b[3]
                               else f"{_e2_slot_b[0]}曜{_e2_slot_b[1]}限（空き）")
                        st.info(f"🅑 コマB: {_lb}")
                    else:
                        _b_hint = "🟢のコマを選択" if _e2_cand_set else "未選択"
                        st.info(f"🅑 コマB: {_b_hint}")

                # 候補ボタン用CSS
                st.markdown("""
<style>
div[data-testid="stMarkdown"]:has(.et-swap-candidate)
  + div[data-testid="stButton"] button {
    background-color: #c8f7c5 !important;
    border: 2px solid #28a745 !important;
    color: #155724 !important;
}
</style>
""", unsafe_allow_html=True)

                # クリック可能な時間割グリッド
                _ppd_e  = st.session_state["periods_per_day"]
                _maxp_e = max(_ppd_e.values()) if _ppd_e else 6

                _hdr_e = st.columns([0.6] + [1] * len(DAYS))
                with _hdr_e[0]:
                    st.markdown("**時限**")
                for _hi_e, _hd_e in enumerate(DAYS):
                    with _hdr_e[_hi_e + 1]:
                        st.markdown(f"**{_hd_e}曜**")

                for _p_e in range(1, _maxp_e + 1):
                    _row_e = st.columns([0.6] + [1] * len(DAYS))
                    with _row_e[0]:
                        st.markdown(f"**{_p_e}限**")
                    for _di_e, _d_e in enumerate(DAYS):
                        with _row_e[_di_e + 1]:
                            if _p_e > _ppd_e.get(_d_e, 6):
                                st.write("─")
                                continue
                            _sc_e = _ss_e = None
                            for _c_e, _sched_e in timetable.items():
                                _s_e = _sched_e.get(_d_e, {}).get(_p_e)
                                if _s_e and edit_teacher in get_teachers_for_slot(_c_e, _s_e):
                                    _sc_e, _ss_e = _c_e, _s_e
                                    break
                            _is_a_e  = bool(_e2_slot_a and _e2_slot_a[0] == _d_e and _e2_slot_a[1] == _p_e)
                            _is_b_e  = bool(_e2_slot_b and _e2_slot_b[0] == _d_e and _e2_slot_b[1] == _p_e)
                            _is_ce   = ((_d_e, _p_e) in _e2_cand_set
                                        and not _is_a_e and not _is_b_e)
                            # 候補セルの前にマーカーdivを挿入
                            if _is_ce:
                                st.markdown(
                                    '<div class="et-swap-candidate"></div>',
                                    unsafe_allow_html=True
                                )
                            _mark_e = "🅐 " if _is_a_e else ("🅑 " if _is_b_e else ("🔄 " if _is_ce else ""))
                            _lbl_e  = f"{_mark_e}{_ss_e}" if _ss_e else f"{_mark_e}空き"
                            if _sc_e:
                                _lbl_e += f"\n({_sc_e})"
                            if st.button(
                                _lbl_e,
                                key=f"et_cell_{_d_e}_{_p_e}",
                                type="primary" if (_is_a_e or _is_b_e) else "secondary",
                                use_container_width=True,
                            ):
                                if _is_a_e:
                                    st.session_state["edit2_slot_a"] = None
                                    st.session_state["edit2_cands"]  = None
                                elif _is_b_e:
                                    st.session_state["edit2_slot_b"] = None
                                elif not _e2_slot_a:
                                    st.session_state["edit2_slot_a"] = (_d_e, _p_e, _sc_e, _ss_e)
                                    st.session_state["edit2_cands"]  = None
                                elif not _e2_slot_b:
                                    st.session_state["edit2_slot_b"] = (_d_e, _p_e, _sc_e, _ss_e)
                                else:
                                    st.session_state["edit2_slot_a"] = (_d_e, _p_e, _sc_e, _ss_e)
                                    st.session_state["edit2_slot_b"] = None
                                    st.session_state["edit2_cands"]  = None
                                st.session_state["edit2_result"] = None
                                st.rerun()

                # A・B 両方選択済みなら即座に結果を表示
                _e2_slot_a = st.session_state.get("edit2_slot_a")
                _e2_slot_b = st.session_state.get("edit2_slot_b")
                if not (_e2_slot_a and _e2_slot_b):
                    st.caption("↑ コマを2つクリックして選択してください")
                else:
                    _e2_da, _e2_pa = _e2_slot_a[0], _e2_slot_a[1]
                    _e2_db, _e2_pb = _e2_slot_b[0], _e2_slot_b[1]
                    _e2_ca, _e2_sa = _e2_slot_a[2], _e2_slot_a[3]
                    _e2_cb, _e2_sb = _e2_slot_b[2], _e2_slot_b[3]

                    _e2_la = (f"【{_e2_ca}】{_e2_da}曜{_e2_pa}限「{_e2_sa}」"
                              if _e2_sa else f"{_e2_da}曜{_e2_pa}限（空きコマ）")
                    _e2_lb = (f"【{_e2_cb}】{_e2_db}曜{_e2_pb}限「{_e2_sb}」"
                              if _e2_sb else f"{_e2_db}曜{_e2_pb}限（空きコマ）")

                    st.markdown("---")
                    if _e2_da == _e2_db and _e2_pa == _e2_pb:
                        st.error("❌ 同じコマが選択されています。")
                    elif not _e2_sa and not _e2_sb:
                        st.warning("⚠️ 両方とも空きコマです。入れ替えの意味がありません。")
                    else:
                        # 実効クラスと検証パラメータを決定
                        if _e2_sa and _e2_sb:
                            _eff_ca, _eff_cb = _e2_ca, _e2_cb
                            _vp_e2 = (_e2_ca, _e2_da, _e2_pa, _e2_cb, _e2_db, _e2_pb)
                        elif _e2_sa:
                            _eff_ca, _eff_cb = _e2_ca, _e2_ca
                            _vp_e2 = (_e2_ca, _e2_da, _e2_pa, _e2_ca, _e2_db, _e2_pb)
                        else:
                            _eff_ca, _eff_cb = _e2_cb, _e2_cb
                            _vp_e2 = (_e2_cb, _e2_db, _e2_pb, _e2_cb, _e2_da, _e2_pa)

                        _ok_e2 = is_valid_swap(timetable, timetable_sg, *_vp_e2)

                        if not _ok_e2:
                            _viols_e2 = get_swap_violations(timetable, timetable_sg, *_vp_e2)
                            st.error("❌ このコマは入れ替えできません。")
                            for _v in _viols_e2:
                                st.write(f"　・{_v}")
                        else:
                            # 他の教員への影響
                            _impacts_e2 = get_swap_impact(
                                timetable, edit_teacher,
                                _eff_ca, _e2_da, _e2_pa,
                                _eff_cb, _e2_db, _e2_pb,
                            )
                            if not _impacts_e2:
                                st.info("他の教員への影響はありません。")
                            else:
                                st.markdown("**他の教員への影響:**")
                                for _imp in _impacts_e2:
                                    st.write(
                                        f"・**{_imp['teacher']}**: "
                                        f"【{_imp['from_cls']}】"
                                        f"{_imp['from_day']}曜{_imp['from_period']}限"
                                        f"「{_imp['from_subj']}」"
                                        f" → "
                                        f"【{_imp['to_cls']}】"
                                        f"{_imp['to_day']}曜{_imp['to_period']}限"
                                    )

                            if st.button("✅ 入れ替えを実行する", type="primary",
                                         key="btn_exec_edit2"):
                                _snap_e2 = {
                                    c: {d: dict(prs) for d, prs in s.items()}
                                    for c, s in timetable.items()
                                }
                                st.session_state["swap_undo_stack"].append({
                                    "snapshot": _snap_e2,
                                    "description": f"{edit_teacher}: {_e2_la} ↔ {_e2_lb}",
                                })
                                if _e2_sa and _e2_sb:
                                    timetable[_e2_ca][_e2_da][_e2_pa] = _e2_sb
                                    timetable[_e2_cb][_e2_db][_e2_pb] = _e2_sa
                                elif _e2_sa:
                                    _ex = timetable.get(_e2_ca, {}).get(_e2_db, {}).get(_e2_pb)
                                    timetable[_e2_ca][_e2_da][_e2_pa] = _ex
                                    timetable[_e2_ca].setdefault(_e2_db, {})[_e2_pb] = _e2_sa
                                else:
                                    _ex = timetable.get(_e2_cb, {}).get(_e2_da, {}).get(_e2_pa)
                                    timetable[_e2_cb][_e2_db][_e2_pb] = _ex
                                    timetable[_e2_cb].setdefault(_e2_da, {})[_e2_pa] = _e2_sb
                                st.session_state["generated_timetable"] = timetable
                                st.session_state["edit2_slot_a"] = None
                                st.session_state["edit2_slot_b"] = None
                                st.session_state["edit2_cands"]  = None
                                st.rerun()

        # ── 元に戻す ──────────────────────────────────────────
        undo_stack = st.session_state.get("swap_undo_stack", [])
        if undo_stack:
            st.markdown("---")
            last = undo_stack[-1]
            st.caption(f"直前の交換: {last['description']}")
            if st.button("↩️ 元に戻す", key="btn_undo_swap"):
                for cls, sched in last["snapshot"].items():
                    for day, periods in sched.items():
                        timetable[cls][day] = dict(periods)
                st.session_state["generated_timetable"] = timetable
                undo_stack.pop()
                st.session_state.pop("swap_candidates", None)
                st.session_state.pop("swap_src", None)
                st.rerun()


# ============================================================
# ⑤ main()（完成版）
# ============================================================

def main():
    st.set_page_config(
        page_title="中学校時間割自動生成",
        layout="wide",
        initial_sidebar_state="expanded"
    )
    init_session_state()
    render_sidebar()

    step = st.session_state["current_step"]
    if   step == 1: render_step1()
    elif step == 2: render_step2()
    elif step == 3: render_step3()
    elif step == 4: render_step4()
    elif step == 5: render_step5()
    elif step == 6: render_step6()
    elif step == 7: render_step7()
    elif step == 8: render_step8()
    elif step == 9: render_step9()
    elif step == 10: render_step10()
    elif step == 0: render_generate()


if __name__ == "__main__":
    main()

